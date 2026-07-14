"""
Proc Aure — Ferramenta de PROCV Inteligente
Desenvolvido por João · Aure Digital
"""

import io
import json
import os
import re
import traceback
from datetime import datetime, timedelta
from typing import Optional, List

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Proc Aure",
    page_icon="✨",
    layout="wide",
    initial_sidebar_state="expanded",  # painel 🤖 Inteligência Claude visível
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

:root{
  --accent:#BF5AF2; --accent-2:#7C5CFF; --accent-3:#E879F9; --accent-green:#30D158;
  --ink:#F5F5F7; --muted:rgba(235,235,245,.5); --muted-2:rgba(235,235,245,.34);
  --glass:rgba(255,255,255,.045); --glass-bd:rgba(255,255,255,.1);
  --ease:cubic-bezier(.16,1,.3,1); --ease-soft:cubic-bezier(.4,0,.2,1);
}

html,body,.stApp,.stMarkdown,p,h1,h2,h3,h4,h5,h6,li,label,button,input,textarea,select,
.stButton,.stTextInput,.stSelectbox,.stMultiSelect,.stMetric{
  font-family:'Inter',-apple-system,BlinkMacSystemFont,'SF Pro Display',sans-serif!important;
  -webkit-font-smoothing:antialiased!important;text-rendering:optimizeLegibility!important}

/* ── Fundo aurora animado ─────────────────────────────────────────── */
.stApp{background:#08070c!important}
.stApp::before{
  content:'';position:fixed;inset:-25%;z-index:0;pointer-events:none;
  background:
    radial-gradient(38% 38% at 18% 22%, rgba(191,90,242,.22), transparent 72%),
    radial-gradient(34% 34% at 82% 14%, rgba(124,92,255,.20), transparent 72%),
    radial-gradient(42% 42% at 68% 86%, rgba(232,121,249,.14), transparent 72%),
    radial-gradient(30% 30% at 30% 78%, rgba(48,209,88,.06), transparent 72%);
  filter:blur(48px);animation:aurora 24s var(--ease-soft) infinite alternate;}
@keyframes aurora{
  0%{transform:translate3d(0,0,0) scale(1) rotate(0deg)}
  50%{transform:translate3d(2.5%,-2%,0) scale(1.1) rotate(2deg)}
  100%{transform:translate3d(-2.5%,2.5%,0) scale(1.05) rotate(-2deg)}}

.main .block-container{position:relative;z-index:1;padding-top:1.5rem!important;
  padding-bottom:6rem!important;max-width:1080px!important}

/* Scrollbar refinada */
::-webkit-scrollbar{width:9px;height:9px}
::-webkit-scrollbar-track{background:transparent}
::-webkit-scrollbar-thumb{background:rgba(191,90,242,.3);border-radius:9px;border:2px solid transparent;background-clip:content-box}
::-webkit-scrollbar-thumb:hover{background:rgba(191,90,242,.55);background-clip:content-box}

/* ── Animações de entrada ─────────────────────────────────────────── */
@keyframes fadeUp{from{opacity:0;transform:translateY(18px)}to{opacity:1;transform:translateY(0)}}
@keyframes shimmer{0%,100%{background-position:0% 50%}50%{background-position:100% 50%}}
@keyframes floatGlow{0%,100%{opacity:.7}50%{opacity:1}}

/* ── Hero ─────────────────────────────────────────────────────────── */
.hero-wrap{text-align:center;padding:2.6rem 0 1rem;animation:fadeUp .9s var(--ease) both}
.hero-badge{display:inline-flex;align-items:center;gap:.5rem;font-size:.66rem;font-weight:600;
  letter-spacing:.2em;text-transform:uppercase;color:rgba(235,235,245,.72);
  border:1px solid rgba(255,255,255,.13);border-radius:999px;padding:.42rem 1.05rem;margin-bottom:1.5rem;
  background:rgba(255,255,255,.05);backdrop-filter:blur(14px);-webkit-backdrop-filter:blur(14px);
  box-shadow:0 2px 20px rgba(0,0,0,.25)}
.hero-badge::before{content:'';width:6px;height:6px;border-radius:50%;
  background:var(--accent-3);box-shadow:0 0 10px var(--accent-3);animation:floatGlow 2.6s ease-in-out infinite}
.hero-title{font-size:4.6rem;font-weight:800;letter-spacing:-.05em;line-height:.96;margin-bottom:1rem;
  background:linear-gradient(115deg,#fff 0%,#F0D9FF 30%,#E879F9 52%,#BF5AF2 72%,#7C5CFF 100%);
  background-size:220% 220%;-webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;
  animation:shimmer 9s ease-in-out infinite;filter:drop-shadow(0 6px 40px rgba(191,90,242,.28))}
.hero-subtitle{color:var(--muted);font-size:1.02rem;font-weight:400;letter-spacing:.005em}

/* ── Passos ───────────────────────────────────────────────────────── */
.step-wrap{display:flex;align-items:center;gap:14px;margin:2.6rem 0 1.4rem;animation:fadeUp .7s var(--ease) both}
.step-num{width:32px;height:32px;border-radius:50%;flex-shrink:0;
  background:linear-gradient(135deg,var(--accent),var(--accent-2));color:#fff;font-size:.78rem;font-weight:700;
  display:flex;align-items:center;justify-content:center;
  box-shadow:0 0 0 1px rgba(255,255,255,.08),0 6px 22px rgba(124,92,255,.5)}
.step-text{font-size:1.12rem;font-weight:600;color:var(--ink);letter-spacing:-.015em}

/* ── Cartões de vidro (uploader, métrica, expander, dataframe) ──────── */
[data-testid='stFileUploader']{border:1px solid var(--glass-bd)!important;border-radius:22px!important;
  background:var(--glass)!important;backdrop-filter:blur(22px)!important;-webkit-backdrop-filter:blur(22px)!important;
  transition:all .45s var(--ease)!important}
[data-testid='stFileUploader']:hover{border-color:rgba(191,90,242,.5)!important;background:rgba(191,90,242,.07)!important;
  transform:translateY(-3px)!important;box-shadow:0 18px 50px rgba(124,92,255,.18)!important}
[data-testid='stFileUploadDropzone']{border:none!important;background:transparent!important}

[data-testid='stMetric']{background:var(--glass)!important;border:1px solid var(--glass-bd)!important;
  border-radius:20px!important;padding:1.3rem 1.5rem!important;backdrop-filter:blur(20px)!important;
  -webkit-backdrop-filter:blur(20px)!important;transition:all .4s var(--ease)!important}
[data-testid='stMetric']:hover{background:rgba(255,255,255,.07)!important;border-color:rgba(191,90,242,.35)!important;
  transform:translateY(-3px)!important;box-shadow:0 18px 44px rgba(0,0,0,.32)!important}
[data-testid='stMetricLabel']{font-size:.7rem!important;font-weight:600!important;text-transform:uppercase!important;
  letter-spacing:.1em!important;color:var(--muted)!important}
[data-testid='stMetricValue']{font-size:2.1rem!important;font-weight:700!important;letter-spacing:-.03em!important;color:var(--ink)!important}
[data-testid='stMetricDelta']{font-size:.82rem!important;font-weight:500!important}

[data-testid='stExpander']{border:1px solid var(--glass-bd)!important;border-radius:18px!important;overflow:hidden!important;
  background:rgba(255,255,255,.03)!important;backdrop-filter:blur(18px)!important;-webkit-backdrop-filter:blur(18px)!important;
  transition:all .35s var(--ease)!important}
[data-testid='stExpander']:hover{border-color:rgba(191,90,242,.28)!important}
[data-testid='stDataFrame']{border-radius:18px!important;overflow:hidden!important;border:1px solid var(--glass-bd)!important}

/* ── Botões ───────────────────────────────────────────────────────── */
[data-testid='stButton']>button{background:linear-gradient(135deg,var(--accent) 0%,var(--accent-2) 100%)!important;
  color:#fff!important;border:none!important;border-radius:16px!important;font-weight:600!important;font-size:1rem!important;
  letter-spacing:.01em!important;padding:.8rem 1.6rem!important;
  box-shadow:0 8px 30px rgba(124,92,255,.38),inset 0 1px 0 rgba(255,255,255,.22)!important;
  transition:all .38s var(--ease)!important}
[data-testid='stButton']>button:hover{transform:translateY(-3px) scale(1.012)!important;
  box-shadow:0 18px 52px rgba(124,92,255,.55),inset 0 1px 0 rgba(255,255,255,.28)!important}
[data-testid='stButton']>button:active{transform:translateY(-1px) scale(.992)!important;transition:all .12s var(--ease)!important}
[data-testid='stDownloadButton']>button{background:linear-gradient(135deg,var(--accent-green) 0%,#1DA844 100%)!important;
  color:#fff!important;border:none!important;border-radius:16px!important;font-weight:600!important;font-size:1rem!important;
  width:100%!important;padding:.9rem!important;
  box-shadow:0 8px 28px rgba(48,209,88,.32),inset 0 1px 0 rgba(255,255,255,.22)!important;
  transition:all .38s var(--ease)!important}
[data-testid='stDownloadButton']>button:hover{transform:translateY(-3px) scale(1.008)!important;
  box-shadow:0 18px 48px rgba(48,209,88,.5)!important}

/* ── Inputs / selects / tags ──────────────────────────────────────── */
[data-baseweb='select']>div{border-radius:13px!important;border-color:rgba(255,255,255,.12)!important;
  background:rgba(255,255,255,.05)!important;backdrop-filter:blur(10px)!important;transition:all .3s var(--ease)!important}
[data-baseweb='select']>div:hover{border-color:rgba(191,90,242,.5)!important;background:rgba(255,255,255,.07)!important}
[data-baseweb='input']{border-radius:13px!important;background:rgba(255,255,255,.05)!important;border-color:rgba(255,255,255,.12)!important}
[data-baseweb='tag']{background:rgba(191,90,242,.2)!important;border:1px solid rgba(191,90,242,.32)!important;border-radius:8px!important}

/* ── Progress / spinner ───────────────────────────────────────────── */
[data-testid='stProgressBar']>div{border-radius:8px!important;background:rgba(255,255,255,.07)!important;height:7px!important}
[data-testid='stProgressBar']>div>div{background:linear-gradient(90deg,var(--accent-3),var(--accent),var(--accent-2))!important;
  border-radius:8px!important;box-shadow:0 0 16px rgba(191,90,242,.5)!important}
.stSpinner>div{border-color:rgba(191,90,242,.15)!important;border-top-color:var(--accent)!important}

/* ── Alertas / divisor / caption ──────────────────────────────────── */
[data-testid='stNotification'],.stAlert{border-radius:16px!important;border-left:none!important;
  backdrop-filter:blur(16px)!important;-webkit-backdrop-filter:blur(16px)!important}
hr{border:none!important;border-top:1px solid rgba(255,255,255,.07)!important;margin:2rem 0!important}
.stCaption,[data-testid='stCaptionContainer']{color:var(--muted-2)!important;font-size:.78rem!important}

/* ── Rodapé ───────────────────────────────────────────────────────── */
.footer{text-align:center;color:rgba(235,235,245,.24);font-size:.76rem;letter-spacing:.05em;
  margin-top:4.5rem;padding-top:1.6rem;border-top:1px solid rgba(255,255,255,.06)}
.footer b{color:rgba(235,235,245,.5);font-weight:600}
</style>""", unsafe_allow_html=True)


# ── Funções de limpeza ─────────────────────────────────────────────────────────

def clean_phone(raw) -> str:
    """
    Trata número com maldade:
    - Remove espaços, traços, pontos, parênteses, barras
    - Remove código do país +55 / 55 quando número tem > 11 dígitos
    - Remove 0 inicial antigo
    - Retorna só dígitos
    """
    if raw is None:
        return ""
    if isinstance(raw, float):
        if np.isnan(raw):
            return ""
        raw = f"{raw:.0f}"  # evita "11987654321.0" e notação científica 6.6e+10
    raw = str(raw).strip()
    if raw.lower() in ("", "nan", "none", "-", "n/a", "#n/a"):
        return ""

    # Número lido como float/científico string: "66999873776.0", "6.6e+10".
    # Só converte quando há '.'/'e' (senão preservaria zeros à esquerda à toa).
    if re.search(r"[.eE]", raw) and re.fullmatch(r"\d+(\.\d+)?([eE][+-]?\d+)?", raw):
        try:
            f = float(raw)
            if f == int(f):
                raw = str(int(f))
        except (ValueError, OverflowError):
            pass

    # Só dígitos (descarta +, texto e símbolos — evita o lixo "66+10")
    digits = re.sub(r"\D", "", raw)

    # Remove DDI 55 se número ficar muito longo
    if digits.startswith("55") and len(digits) > 11:
        digits = digits[2:]

    # Remove 0 de discagem longa (ex: 011...)
    if digits.startswith("0") and len(digits) >= 11:
        digits = digits[1:]

    return digits


def right8(phone_clean: str) -> str:
    """Últimos 8 dígitos — padroniza números com/sem 9 e com/sem DDD.
    Retorna '' se < 8 dígitos ou se for placeholder (todos os dígitos iguais)."""
    if not phone_clean:
        return ""
    digits = re.sub(r"\D", "", str(phone_clean))
    if len(digits) < 8:
        return ""  # Muito curto para ser telefone — descarta
    sub8 = digits[-8:]
    if len(set(sub8)) <= 1:
        return ""  # 00000000, 99999999… — placeholder, não é telefone real
    return sub8


# DDDs brasileiros oficiais — usado para separar telefone de CPF/CEP/ID.
_VALID_DDDS = {str(d) for d in (
    11, 12, 13, 14, 15, 16, 17, 18, 19,
    21, 22, 24, 27, 28,
    31, 32, 33, 34, 35, 37, 38,
    41, 42, 43, 44, 45, 46, 47, 48, 49,
    51, 53, 54, 55,
    61, 62, 63, 64, 65, 66, 67, 68, 69,
    71, 73, 74, 75, 77, 79,
    81, 82, 83, 84, 85, 86, 87, 88, 89,
    91, 92, 93, 94, 95, 96, 97, 98, 99,
)}


def _looks_like_phone(value) -> bool:
    """True se o valor parece um telefone brasileiro real.
    Rejeita datas, dígitos repetidos, CPF/CEP e comprimentos fora da faixa.
    - 11 dígitos → celular: exige DDD válido + 3º dígito '9'
    - 10 dígitos → fixo:    exige DDD válido
    - 8-9 dígitos → telefone sem DDD (aceito, sinal mais fraco)
    """
    s = str(value).strip()
    if not s:
        return False
    digits = clean_phone(value)
    n = len(digits)
    if n < 8 or n > 11:
        return False
    if len(set(digits[-8:])) <= 1:  # placeholder
        return False
    # Data disfarçada ("2024-01-15", "15/03/2024") não é telefone
    if re.search(r"[/\-]", s) and parse_date(value) is not None:
        return False
    if n == 11:
        return digits[:2] in _VALID_DDDS and digits[2] == "9"
    if n == 10:
        return digits[:2] in _VALID_DDDS
    return True  # 8-9 dígitos


def phone_key(raw) -> tuple[Optional[str], str]:
    """Chave canônica de telefone: (ddd, sub8).
    sub8 = 8 últimos dígitos (casa com/sem o 9 e com/sem DDD).
    ddd  = 2 primeiros dígitos quando há 10-11 dígitos com DDD válido; senão None.
    Retorna (None, '') para placeholders / números curtos demais."""
    digits = clean_phone(raw)
    sub8 = right8(digits)
    if not sub8:
        return (None, "")
    ddd = None
    if len(digits) in (10, 11):
        cand = digits[:2]
        if cand in _VALID_DDDS:
            ddd = cand
    return (ddd, sub8)


def phone_group_key(raw) -> str:
    """Chave de agrupamento de telefone para dedup: DDD+sub8 quando há DDD, senão sub8.
    Mantém pessoas de DDDs diferentes em grupos diferentes."""
    ddd, sub8 = phone_key(raw)
    if not sub8:
        return ""
    return f"{ddd}{sub8}" if ddd else sub8


def _row_ident(ddd: Optional[str], sub8: str, primary_raw="") -> str:
    """Identidade do comprador (DDD+sub8) para dedup e sobreposição tráfego×disparo.
    Usa a chave que de fato casou; se o match foi por nome (sem telefone), usa o telefone principal."""
    if sub8:
        return f"{ddd}{sub8}" if ddd else sub8
    return phone_group_key(primary_raw)


def phones_match(a: tuple, b: tuple) -> bool:
    """True se duas chaves casam: sub8 igual E DDD compatível (igual ou ausente num lado).
    É o que mata o falso match entre cidades diferentes — (11) x (21) com mesmo final."""
    ddd_a, s_a = a
    ddd_b, s_b = b
    if not s_a or s_a != s_b:
        return False
    if ddd_a and ddd_b and ddd_a != ddd_b:
        return False
    return True


def add_to_phone_lookup(lookup: dict, key: tuple, payload) -> None:
    """Adiciona payload ao lookup {sub8: [(ddd, payload), ...]} preservando ordem de prioridade."""
    ddd, sub8 = key
    if not sub8:
        return
    lookup.setdefault(sub8, []).append((ddd, payload))


def resolve_phone(lookup: dict, key: tuple):
    """Retorna o payload com DDD compatível (prefere DDD exato; senão lado-ausente).
    Entre candidatos de mesma compatibilidade, vence o primeiro inserido (maior prioridade)."""
    ddd, sub8 = key
    if not sub8:
        return None
    candidates = lookup.get(sub8)
    if not candidates:
        return None
    if ddd:  # 1) DDD exato
        for c_ddd, payload in candidates:
            if c_ddd == ddd:
                return payload
    for c_ddd, payload in candidates:  # 2) lado sem DDD (compatível)
        if c_ddd is None or ddd is None:
            return payload
    return None  # 3) só há DDDs diferentes — não casa


def resolve_phone_all(lookup: dict, key: tuple) -> list:
    """Todos os payloads com DDD compatível (DDD exato primeiro, depois lado-ausente).
    Usado no disparo, onde um mesmo número pode ter várias vendas (multi-compra)."""
    ddd, sub8 = key
    if not sub8:
        return []
    candidates = lookup.get(sub8)
    if not candidates:
        return []
    exact = [p for c_ddd, p in candidates if ddd and c_ddd == ddd]
    loose = [p for c_ddd, p in candidates if c_ddd is None or ddd is None]
    return exact + loose


# ── Detecção automática de colunas ─────────────────────────────────────────────

def _match_keywords(col_name: str, keywords: list[str]) -> bool:
    c = col_name.lower().strip()
    return any(kw in c for kw in keywords)


@st.cache_data(show_spinner=False)
def detect_phone_col(df: pd.DataFrame) -> Optional[str]:
    """
    Detecta coluna de telefone por conteúdo (7+ dígitos) + bônus por nome.
    Funciona com colunas sem nome óbvio — cada planilha é um caso.
    """
    kws = [
        "telefone", "celular", "whatsapp", "wpp", "fone", "phone",
        "mobile", "cel", "numero", "número", "tel", "contato",
        "phone_number", "nr_tel", "nro", "n_tel",
    ]
    # "id" com fronteira de palavra — substring pegaria "Telefone resIDencial"
    skip_kws = ["data", "date", "valor", "preco", "preço", "total",
                "cpf", "cnpj", "cep", "código", "codigo",
                "documento", "doc", "inscricao", "inscrição"]
    _id_re = re.compile(r"(?:^|[^a-zà-ÿ])id(?:[^a-zà-ÿ]|$)")

    def _phone_hits(col: str) -> int:
        sample = df[col].dropna().astype(str).head(30)
        if len(sample) == 0:
            return 0
        return int(sample.apply(_looks_like_phone).sum())

    candidates = []
    for col in df.columns:
        if str(col).startswith("_") or _match_keywords(col, skip_kws) or _id_re.search(str(col).lower()):
            continue
        hits = _phone_hits(col)
        if hits < 2:
            continue
        sample = df[col].dropna().astype(str).head(30)
        sample_size = max(1, len(sample))
        content_score = (hits / sample_size) * 60
        name_bonus = 40 if _match_keywords(col, kws) else 0
        # Cobertura desempata (contínuo, sem truncar): 'Celular 1' com 50
        # preenchidos vence 'Telefone' com 48.
        coverage_bonus = len(df[col].dropna()) / max(1, len(df)) * 15
        # Celular > fixo: leads de CRM/WhatsApp casam por celular. Conta os
        # números com o 9 de celular (11 díg. com DDD ou 9 sem DDD).
        mobile_hits = sum(
            1 for v in sample
            for d in [clean_phone(v)]
            if (len(d) == 11 and d[2] == "9") or (len(d) == 9 and d[0] == "9")
        )
        mobile_bonus = (mobile_hits / sample_size) * 8
        candidates.append((col, content_score + name_bonus + coverage_bonus + mobile_bonus))

    if not candidates:
        return None
    return max(candidates, key=lambda x: x[1])[0]


@st.cache_data(show_spinner=False)
def detect_tag_col(df: pd.DataFrame) -> Optional[str]:
    kws = ["tag", "etiqueta", "label", "categoria", "tipo", "fonte", "origem"]
    for col in df.columns:
        if str(col).startswith("_"):
            continue
        if _match_keywords(col, kws):
            return col
    return None


@st.cache_data(show_spinner=False)
def detect_name_col(df: pd.DataFrame) -> Optional[str]:
    """Ordem dos kws = prioridade ('Nome do contato' vence 'ID contato').
    Valida o conteúdo: coluna de nome precisa ter texto, não números/IDs."""
    kws = ["nome", "name", "cliente", "comprador", "titular", "razao", "razão", "social", "contato"]
    skip_kws = ["telefone", "cel", "phone", "whatsapp", "wpp", "email",
                "cidade", "estado", "produto", "tag", "etiqueta"]

    def _has_text_content(col: str) -> bool:
        sample = df[col].dropna().astype(str).head(20)
        if len(sample) == 0:
            return False
        with_letters = sum(1 for v in sample if len(re.findall(r"[^\W\d_]", v)) >= 2)
        return with_letters / len(sample) >= 0.6

    for kw in kws:
        for col in df.columns:
            if str(col).startswith("_") or _match_keywords(col, skip_kws):
                continue
            if kw in str(col).lower() and _has_text_content(col):
                return col
    return None


@st.cache_data(show_spinner=False)
def detect_value_col(df: pd.DataFrame) -> Optional[str]:
    """
    Detecta coluna de valor monetário por análise estatística — sem depender de keywords.
    Uma coluna de preço tem: separadores decimais, valores não-zero, faixa razoável, variância.
    Keywords de nome dão um bônus pequeno mas não são obrigatórias.
    """
    # Colunas que NUNCA são preço — só por nome óbvio
    hard_skip = ["cpf", "cnpj", "rg", "tel", "cel", "fone", "phone",
                 "whatsapp", "wpp", "percentual", "percent", "taxa", "mrg",
                 "margem", "desconto", "devolucao", "devolução"]

    # Keywords de nome para bônus (não obrigatório)
    bonus_kws = ["valor", "value", "total", "preco", "preço", "receita",
                 "ticket", "fatura", "price", "vl", "vlr", "bruto", "liquido",
                 "líquido", "sale", "compra", "amount"]

    def _parse_price(v):
        """Retorna float se o valor parece preço, None caso contrário."""
        s = re.sub(r"[R$\s]", "", str(v).strip())
        if not s or not re.search(r"\d", s):
            return None
        if re.search(r"[A-Za-z]", s):
            return None  # "30-Jun-26", códigos etc. — não é preço
        if "," in s:
            try:
                val = float(s.replace(".", "").replace(",", "."))
                return val if val < 10_000_000 else None
            except ValueError:
                return None
        # Decimal americano ("8386.5", "3669.68000000000") — exatamente 3 casas
        # é milhar BR ("1.234"), que cai no caminho de dígitos puros abaixo.
        m_us = re.fullmatch(r"(\d+)\.(\d+)", s)
        if m_us and len(m_us.group(2)) != 3:
            try:
                val = float(s)
                return val if val < 10_000_000 else None
            except ValueError:
                return None
        # Sem decimal: só aceita se < 8 dígitos (afasta CPF/tel/CNPJ)
        pure = re.sub(r"\D", "", s)
        if 0 < len(pure) < 8:
            try:
                return float(pure)
            except ValueError:
                return None
        return None

    candidates: list = []
    for col in df.columns:
        col_lower = str(col).lower()
        if str(col).startswith("_") or any(kw in col_lower for kw in hard_skip) or _is_doc_col(col):
            continue

        sample = df[col].dropna().head(50)
        if len(sample) < 3:
            continue

        parsed = [_parse_price(v) for v in sample]
        vals = [v for v in parsed if v is not None]
        if len(vals) < max(2, len(sample) * 0.35):
            continue

        n_nonzero = sum(1 for v in vals if v > 0)
        n_decimal = sum(
            1 for v, p in zip(sample, parsed)
            if "," in str(v) or (p is not None and p != int(p))
        )
        mean_val = sum(vals) / len(vals)

        # Coeficiente de variação: preços reais têm variância razoável
        # (não são todos iguais como "01", "01", "01")
        if mean_val > 0:
            std_val = (sum((v - mean_val) ** 2 for v in vals) / len(vals)) ** 0.5
            cv = std_val / mean_val
        else:
            cv = 0

        # Pontuação estatística — sem depender de nome
        score = 0
        score += int((n_decimal / len(sample)) * 35)          # decimal → forte sinal
        score += int((n_nonzero / max(len(vals), 1)) * 30)    # não-zero → valor real
        score += 20 if 1 <= mean_val <= 500_000 else 0        # faixa de preço razoável
        score += 10 if 0.05 <= cv <= 8.0 else 0               # variância adequada
        score -= 15 if cv < 0.01 and len(vals) > 3 else 0     # penaliza coluna uniforme

        # Bônus por nome (pequeno — não é obrigatório)
        if any(kw in col_lower for kw in bonus_kws):
            score += 15

        if score > 10:
            candidates.append((col, score))

    if not candidates:
        return None
    return max(candidates, key=lambda x: x[1])[0]


def parse_value(v) -> float:
    """Converte string de valor monetário para float, tolerando:
    R$, espaços, milhar/decimal em formato BR (1.234,56) ou US (1,234.56),
    negativos com sinal (-50) ou contábil ((50,00)) e ruído de texto."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    neg = (s.startswith("(") and s.endswith(")")) or s.lstrip().startswith("-")
    s = re.sub(r"[^\d.,]", "", s)  # mantém só dígitos, ponto e vírgula
    if not s:
        return 0.0
    last_c, last_d = s.rfind(","), s.rfind(".")
    if last_c >= 0 and last_d >= 0:
        if last_c > last_d:                          # vírgula decimal (BR): 1.234,56
            s = s.replace(".", "").replace(",", ".")
        else:                                        # ponto decimal (US): 1,234.56
            s = s.replace(",", "")
    elif last_c >= 0:                                # só vírgula → decimal BR
        s = s.replace(".", "").replace(",", ".")
    elif "." in s and re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
        s = s.replace(".", "")                       # ponto como milhar BR: 2.500
    try:
        val = float(s)
    except ValueError:
        return 0.0
    return -val if neg else val


# ── Carregamento de arquivo ────────────────────────────────────────────────────

class _FileLike:
    """Wrapper de bytes para simular um arquivo uploadado (necessário para cache)."""
    def __init__(self, data: bytes, name: str):
        self._buf = io.BytesIO(data)
        self.name = name

    def __getattr__(self, attr):
        """Delega QUALQUER método não definido aqui para o BytesIO interno.
        Isso cobre tell, seekable, readable, getvalue, etc. sem precisar
        declarar cada um individualmente."""
        return getattr(self._buf, attr)

    def read(self, *a): return self._buf.read(*a)
    def seek(self, *a): return self._buf.seek(*a)
    def tell(self, *a): return self._buf.tell(*a)


@st.cache_data(show_spinner=False)
def _load_cached(file_bytes: bytes, file_name: str, sheets: tuple):
    """Versão cacheada do load_file_multisheet — evita re-parsear no mesmo arquivo."""
    return load_file_multisheet(_FileLike(file_bytes, file_name), list(sheets))


@st.cache_data(show_spinner=False)
def _get_sheets_cached(file_bytes: bytes, file_name: str) -> list:
    return get_excel_sheets(_FileLike(file_bytes, file_name))


def get_excel_sheets(uploaded) -> list:
    """Retorna lista de abas de um Excel. Lista vazia indica CSV."""
    name = uploaded.name.lower()
    if not name.endswith((".xlsx", ".xls", ".xlsm")):
        return []
    try:
        return _open_excel_file(uploaded).sheet_names
    except Exception:
        return []


def _looks_like_data_row(row) -> bool:
    """True se a linha parece DADO (telefone/data/valor) e não rótulos de cabeçalho."""
    non_null = [v for v in row if pd.notna(v) and str(v).strip().lower() not in ("", "nan")]
    if not non_null:
        return False
    data_hits = sum(
        1 for v in non_null
        if _looks_like_phone(v)
        or parse_date(v) is not None
        or re.fullmatch(r"[\d\s\.\,\-\+\/R$%]+", str(v).strip()) is not None
    )
    return data_hits / len(non_null) >= 0.5


def detect_header_row(df_raw: pd.DataFrame) -> int:
    """
    Varre as primeiras linhas e retorna o índice da linha que parece
    ser o cabeçalho real (ignora títulos e linhas em branco acima).
    Retorna -1 quando a planilha NÃO tem cabeçalho (dados já na 1ª linha preenchida).
    """
    n_cols = len(df_raw.columns)
    scan = min(30, len(df_raw))

    # Preenchimento por linha. Relatórios de ERP têm MUITAS colunas vazias e o
    # cabeçalho usa poucas células — o piso é adaptativo à linha mais densa.
    fills = []
    for i in range(scan):
        row = df_raw.iloc[i]
        fills.append(sum(1 for v in row if pd.notna(v) and str(v).strip() not in ("", "nan")))
    densest = max(fills, default=0)
    min_filled = max(2, min(int(n_cols * 0.3), max(2, int(densest * 0.5))))

    for i in range(scan):
        if fills[i] < min_filled:
            continue
        row = df_raw.iloc[i]
        non_null = [v for v in row if pd.notna(v) and str(v).strip() not in ("", "nan")]
        label_count = sum(
            1 for v in non_null
            if isinstance(v, str)
            and not re.fullmatch(r"[\d\s\.\,\-\+\/\%]+", v.strip())
        )
        # Rótulos suficientes E a linha não é dado disfarçado → é o cabeçalho
        if label_count / len(non_null) >= 0.4 and not _looks_like_data_row(row):
            return i

    # Nenhum cabeçalho claro: se a 1ª linha preenchida já é dado, é planilha SEM cabeçalho
    for i in range(scan):
        if fills[i] >= min_filled:
            return -1 if _looks_like_data_row(df_raw.iloc[i]) else 0
    return 0


def _is_blank_col(series: pd.Series) -> bool:
    """True se a coluna está inteiramente vazia (NaN ou string vazia)."""
    s = series.astype(str).str.strip().str.lower()
    return bool(((series.isna()) | (s == "") | (s == "nan")).all())


def _split_side_by_side_blocks(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Detecta planilha dividida em blocos lado a lado (separados por coluna vazia)
    e os reempilha verticalmente num só bloco. Conservador: só junta blocos de
    MESMA largura, cada um com ao menos uma coluna de telefone. Senão, devolve igual.
    Recebe e devolve o DataFrame bruto (header=None).
    """
    n = len(df_raw.columns)
    if n < 4 or len(df_raw) < 2:
        return df_raw

    empty_cols = {i for i in range(n) if _is_blank_col(df_raw.iloc[:, i])}
    if not empty_cols:
        return df_raw

    # Quebra em blocos contíguos de colunas não-vazias
    blocks, cur = [], []
    for i in range(n):
        if i in empty_cols:
            if cur:
                blocks.append(cur); cur = []
        else:
            cur.append(i)
    if cur:
        blocks.append(cur)

    if len(blocks) < 2 or len({len(b) for b in blocks}) != 1:
        return df_raw  # menos de 2 blocos ou larguras diferentes → não é a mesma tabela

    def _is_header_like(vals) -> bool:
        nn = [v for v in vals if pd.notna(v) and str(v).strip().lower() not in ("", "nan")]
        if not nn:
            return False
        labels = sum(1 for v in nn if not re.fullmatch(r"[\d\s\.\,\-\+\/R$%]+", str(v).strip()))
        return labels / len(nn) >= 0.5

    per_block_header = all(_is_header_like(df_raw.iloc[0, b].tolist()) for b in blocks)

    sub_frames = []
    for bi, b in enumerate(blocks):
        sub = df_raw.iloc[:, b].reset_index(drop=True)
        if not any(_is_phone_col(sub.iloc[:, j].dropna()) for j in range(sub.shape[1])):
            return df_raw  # bloco sem telefone → não arrisca juntar tabelas diferentes
        if per_block_header and bi > 0:
            sub = sub.iloc[1:].reset_index(drop=True)  # remove cabeçalho repetido do bloco
        sub.columns = range(sub.shape[1])
        sub_frames.append(sub)

    stacked = pd.concat(sub_frames, ignore_index=True)
    stacked.columns = range(stacked.shape[1])
    stacked.attrs["blocks_merged"] = len(blocks)
    return stacked


_SUMMARY_WORDS = {"total", "totais", "subtotal", "soma", "somatorio", "geral", "media", "media:"}


def _drop_summary_rows(df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
    """Remove linhas de fechamento (TOTAL/SOMA/SUBTOTAL) que não são vendas.
    Conservador: só dropa se a 1ª palavra de alguma célula é de soma E a linha não tem telefone."""
    if len(df) == 0:
        return df, 0

    def _is_summary(row) -> bool:
        cells = [v for v in row if pd.notna(v) and str(v).strip() != ""]
        if not cells or any(_looks_like_phone(v) for v in cells):
            return False
        for v in cells:
            words = re.sub(r"[^a-z ]", " ", _normalize(str(v))).split()
            if words and words[0] in _SUMMARY_WORDS:
                return True
        return False

    mask = df.apply(_is_summary, axis=1)
    n = int(mask.sum())
    if n:
        return df[~mask].reset_index(drop=True), n
    return df, 0


def _frame_from_raw(df_raw: pd.DataFrame) -> tuple[pd.DataFrame, int, list]:
    """
    Recebe o DataFrame bruto (lido com header=None) e devolve (df tratado, hrow, notas).
    Trata: blocos lado a lado, detecção de cabeçalho, planilha sem cabeçalho e linhas de soma.
    """
    notes: list = []
    if len(df_raw) == 0 or len(df_raw.columns) == 0:
        return pd.DataFrame(), 0, []
    df_raw = _split_side_by_side_blocks(df_raw)
    blocks_merged = df_raw.attrs.get("blocks_merged", 0)
    if blocks_merged:
        notes.append(f"Aba dividida em {blocks_merged} blocos lado a lado — reempilhados num só.")

    hrow = detect_header_row(df_raw)
    if hrow == -1:
        df = df_raw.copy()
        df.columns = [f"Coluna {i + 1}" for i in range(len(df.columns))]
        notes.append("Planilha sem cabeçalho — colunas nomeadas como Coluna 1, 2, 3…")
    else:
        header_vals = df_raw.iloc[hrow].tolist()
        df = df_raw.iloc[hrow + 1:].copy()
        df.columns = [
            str(v).strip() if pd.notna(v) and str(v).strip().lower() not in ("", "nan")
            else f"Coluna {i + 1}"
            for i, v in enumerate(header_vals)
        ]
    df = df.dropna(how="all").reset_index(drop=True)
    df, n_sum = _drop_summary_rows(df)
    if n_sum:
        notes.append(f"{n_sum} linha(s) de total/soma ignorada(s) (não são vendas).")
    return df, hrow, notes


def _load_single_sheet(xls, sheet_name: str) -> tuple[pd.DataFrame, int, list]:
    df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, dtype=str)
    return _frame_from_raw(df_raw)


def _open_excel_file(uploaded):
    """Tenta abrir um ExcelFile com múltiplos engines (openpyxl, xlrd)."""
    for engine in (None, "openpyxl", "xlrd"):
        try:
            uploaded.seek(0)
            if engine:
                return pd.ExcelFile(uploaded, engine=engine)
            return pd.ExcelFile(uploaded)
        except Exception:
            continue
    raise ValueError("Não foi possível abrir o arquivo Excel. Verifique se o arquivo não está corrompido.")


def _load_csv_smart(uploaded) -> tuple[pd.DataFrame, int, list]:
    """Lê CSV detectando encoding E separador (vírgula, ponto-e-vírgula, tab, pipe)."""
    uploaded.seek(0)
    raw = uploaded.read()
    if isinstance(raw, str):
        raw = raw.encode("utf-8")
    for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252", "iso-8859-1"):
        try:
            text = raw.decode(enc)
        except (UnicodeDecodeError, LookupError):
            continue
        # Detecta o separador pela 1ª linha não-vazia (mais frequente entre os candidatos)
        first = next((ln for ln in text.splitlines() if ln.strip()), "")
        delim = max((",", ";", "\t", "|"), key=first.count)
        if first.count(delim) == 0:
            delim = ","
        try:
            df_raw = pd.read_csv(io.StringIO(text), header=None, dtype=str, sep=delim)
            return _frame_from_raw(df_raw)
        except Exception:
            continue
    return pd.DataFrame(), 0, []


def _score_sheet(df: pd.DataFrame) -> int:
    """
    Pontua uma aba como planilha de vendas (0-100).
    Abas primárias têm: data, valor, telefone e poucos 'Unnamed'.
    """
    score = 0
    if detect_date_col(df) is not None:
        score += 35   # data é o sinal mais forte
    if detect_value_col(df) is not None:
        score += 30
    if detect_phone_col(df) is not None:
        score += 20
    unnamed = sum(1 for c in df.columns if str(c).lower().startswith("unnamed"))
    unnamed_ratio = unnamed / max(len(df.columns), 1)
    if unnamed_ratio < 0.2:
        score += 15
    return score


def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Apara pontas, colapsa espaços internos e resolve colunas duplicadas."""
    cols = [re.sub(r"\s+", " ", str(c).strip()) for c in df.columns]
    seen: dict = {}
    clean = []
    for c in cols:
        if c in seen:
            seen[c] += 1
            clean.append(f"{c}_{seen[c]}")
        else:
            seen[c] = 0
            clean.append(c)
    df = df.copy()
    df.columns = clean
    return df


def _canon_colname(c: str) -> str:
    """Nome canônico de coluna: minúsculas, sem acento, espaços colapsados."""
    return _normalize(re.sub(r"\s+", " ", str(c).strip()))


def _align_columns(dfs: list) -> tuple[list, int]:
    """
    Renomeia colunas equivalentes (mesmo nome canônico) entre abas para o MESMO
    rótulo, evitando que 'Telefone' e 'Telefone ' virem colunas separadas no concat.
    Devolve (lista de DataFrames renomeados, nº de ajustes).
    """
    if len(dfs) < 2:
        return dfs, 0
    canon_to_display: dict = {}
    for df in dfs:
        for c in df.columns:
            k = _canon_colname(c)
            if k and k not in canon_to_display:
                canon_to_display[k] = c
    n_aligned = 0
    out = []
    for df in dfs:
        rename = {}
        for c in df.columns:
            target = canon_to_display.get(_canon_colname(c), c)
            if target != c and target not in df.columns:
                rename[c] = target
                n_aligned += 1
        out.append(df.rename(columns=rename) if rename else df)
    return out, n_aligned


def load_file_multisheet(
    uploaded, selected_sheets: list
) -> tuple[Optional[pd.DataFrame], dict]:
    """
    Carrega uma ou mais abas inteligentemente.
    Quando há múltiplas abas, pontua cada uma como planilha de vendas.
    Abas primárias (score ≥ 50) vêm primeiro no DataFrame combinado — isso
    garante que o lookup de telefone prefira linhas com data e valor.
    CSV: selected_sheets ignorado.
    """
    name = uploaded.name.lower()
    sheet_info: dict = {}
    scored_dfs: list = []   # (score, sheet_name, df, hrow)
    treatment_notes: list = []

    try:
        if name.endswith(".csv"):
            df, hrow, notes = _load_csv_smart(uploaded)
            if len(df) > 0:
                df = _normalize_cols(df)
                sheet_info["CSV"] = hrow
                scored_dfs.append((100, "CSV", df, hrow))
                treatment_notes.extend(notes)
        elif name.endswith((".xlsx", ".xls", ".xlsm")):
            xls = _open_excel_file(uploaded)
            for sheet in selected_sheets:
                try:
                    df, hrow, notes = _load_single_sheet(xls, sheet)
                except Exception as e:
                    treatment_notes.append(
                        f"Aba '{sheet}' ignorada — erro ao ler ({type(e).__name__}): {e}"
                    )
                    continue
                if len(df) > 0:
                    df = _normalize_cols(df)
                    score = _score_sheet(df)
                    sheet_info[sheet] = hrow
                    scored_dfs.append((score, sheet, df, hrow))
                    treatment_notes.extend(
                        (f"Aba '{sheet}': {n}" if len(selected_sheets) > 1 else n)
                        for n in notes
                    )
    except Exception as e:
        st.error(f"Erro ao ler arquivo: {e}")
        return None, {}

    if not scored_dfs:
        return None, {}

    # Ordena: abas primárias (maior score) primeiro
    # Isso faz com que o lookup de telefone prefira linhas com data+valor
    scored_dfs.sort(key=lambda x: x[0], reverse=True)

    # Alinha colunas equivalentes entre abas ANTES do concat (corrige fragmentação)
    aligned_dfs, n_aligned = _align_columns([t[2] for t in scored_dfs])
    if n_aligned > 0:
        treatment_notes.append(
            f"Colunas equivalentes entre abas foram unificadas ({n_aligned} ajuste(s)) — "
            f"evita dados fragmentados (ex.: 'Telefone' e 'Telefone ')."
        )

    dfs = []
    primary_sheets = []
    secondary_sheets = []
    for (score, sheet_name, _orig, hrow), df in zip(scored_dfs, aligned_dfs):
        if len(scored_dfs) > 1:
            df = df.copy()
            df.insert(0, "_Planilha", sheet_name)
            df.insert(1, "_Score_Aba", score)
        if score >= 50:
            primary_sheets.append(sheet_name)
        else:
            secondary_sheets.append(sheet_name)
        dfs.append(df)

    combined = _normalize_cols(pd.concat(dfs, ignore_index=True))

    # Avisos para a UI
    if secondary_sheets:
        combined.attrs["secondary_sheets"] = secondary_sheets
        combined.attrs["primary_sheets"] = primary_sheets
    if treatment_notes:
        combined.attrs["treatment_notes"] = treatment_notes

    return combined, sheet_info


def _unify_key_cols(frames: list, detector) -> list:
    """Quando funis nomeiam a MESMA coluna-chave de formas diferentes (ex.: 'Celular'
    num funil e 'Telefone' noutro), renomeia todas para um nome comum, para que o
    cruzamento enxergue a coluna em todos. Não mexe quando já estão consistentes."""
    detected = [detector(d) for d in frames]
    present = [c for c in detected if c]
    if len(set(present)) <= 1:
        return frames  # todos iguais (ou nenhum) — nada a unificar
    target = present[0]  # usa o nome do primeiro funil que tem a coluna
    out = []
    for d, col in zip(frames, detected):
        if col and col != target and target not in d.columns:
            d = d.rename(columns={col: target})
        out.append(d)
    return out


def combine_kommo_sources(dfs: list, names: Optional[list] = None) -> Optional[pd.DataFrame]:
    """
    Combina vários exports do Kommo (ex.: um por funil de venda) num só DataFrame.
    Unifica as colunas de telefone e de tags (nomes podem variar entre funis),
    alinha o resto e marca a origem em '_Funil'.

    Um lead pode aparecer em mais de um funil; telefones de QUALQUER funil entram
    no cruzamento e a deduplicação por comprador acontece depois (no run_procv),
    então não há dupla contagem de conversões.
    """
    pairs = [(d, (names[i] if names else f"Funil {i+1}"))
             for i, d in enumerate(dfs) if d is not None and len(d) > 0]
    if not pairs:
        return None
    if len(pairs) == 1:
        return pairs[0][0]

    frames = [d for d, _ in pairs]
    frames = _unify_key_cols(frames, detect_phone_col)
    frames = _unify_key_cols(frames, detect_tag_col)
    aligned, _ = _align_columns(frames)
    out = []
    notes = []
    for (d_orig, nm), d in zip(pairs, aligned):
        d = d.copy()
        d.insert(0, "_Funil", nm)
        out.append(d)
        notes.append(f"{nm}: {len(d):,} leads")
    combined = _normalize_cols(pd.concat(out, ignore_index=True))
    combined.attrs["treatment_notes"] = (
        [f"{len(pairs)} arquivos do Kommo combinados (funis): " + " · ".join(notes)]
    )
    return combined


def combine_sales_sources(dfs: list, names: Optional[list] = None) -> Optional[pd.DataFrame]:
    """
    Combina várias LISTAS DE VENDAS (ex.: uma por vendedora) num só DataFrame.

    Cada lista vira um grupo na coluna '_Planilha' — o mesmo mecanismo do
    breakdown mês a mês passa a entregar o resultado SEPARADO POR LISTA, e o
    cruzamento principal entrega o total combinado (comprador repetido entre
    listas conta uma vez). Colunas-chave com nomes diferentes entre as listas
    (ex.: 'Telefone' numa, 'Celular' noutra) são unificadas.
    """
    pairs = [(d, (names[i] if names else f"Lista {i + 1}"))
             for i, d in enumerate(dfs) if d is not None and len(d) > 0]
    if not pairs:
        return None
    if len(pairs) == 1:
        return pairs[0][0]

    frames, stems, all_notes = [], [], []
    for d, nm in pairs:
        stem = re.sub(r"\.(xlsx|xlsm|xls|csv)$", "", str(nm), flags=re.I)
        stems.append(stem)
        all_notes.extend(f"'{stem}': {n}" for n in d.attrs.get("treatment_notes", []))
        d = d.copy()
        if "_Planilha" in d.columns:  # arquivo multi-aba: prefixa a aba com a lista
            d["_Planilha"] = stem + " · " + d["_Planilha"].astype(str)
        else:
            d.insert(0, "_Planilha", stem)
        if "_Score_Aba" not in d.columns:
            d.insert(1, "_Score_Aba", 100)  # lista de vendas é sempre primária
        frames.append(d)

    for detector in (detect_phone_col, detect_name_col, detect_value_col, detect_date_col):
        frames = _unify_key_cols(frames, detector)
    aligned, _ = _align_columns(frames)
    combined = _normalize_cols(pd.concat(aligned, ignore_index=True))
    resumo = " · ".join(f"{s}: {len(d):,} vendas" for s, (d, _) in zip(stems, pairs))
    combined.attrs["treatment_notes"] = (
        [f"{len(pairs)} listas de vendas combinadas — {resumo}. "
         f"O resultado sai por lista E com o total junto (comprador repetido entre listas conta 1× no total)."]
        + all_notes
    )
    return combined


# ── Lógica principal do PROCV ──────────────────────────────────────────────────

def _is_phone_col(col_data: pd.Series) -> bool:
    """Retorna True se a coluna tem pelo menos 2 valores que parecem telefone real.
    Amostra os primeiros 50 valores não-nulos — rápido e suficiente para classificar."""
    if len(col_data) == 0:
        return False
    sample = col_data.head(50)
    phone_like = int(sample.apply(_looks_like_phone).sum())
    return phone_like >= 2


_DOC_COLS = {"cpf", "cnpj", "rg", "documento", "doc", "inscricao", "inscrição",
             "cadastro_pessoa", "cpf_cnpj", "identificacao", "identificação"}

def _is_doc_col(col_name: str) -> bool:
    """Retorna True se o nome da coluna sugere CPF/CNPJ/RG — evita falsos matches."""
    c = col_name.lower()
    return any(kw in c for kw in _DOC_COLS)


# Token de telefone BR embutido em texto: (DDI) (DDD) 9? XXXX(-)XXXX
# Token de telefone BR embutido em texto. O "(?:9[\s.\-]?)?" cobre o 9 do celular
# isolado por separador — ex.: "(11) 9 8888-7777" / "11 9 8888-7777".
_PHONE_TOKEN_RE = re.compile(r"(?:\+?55[\s.\-]?)?\(?\d{2}\)?[\s.\-]?(?:9[\s.\-]?)?\d{4}[\s.\-]?\d{4}")


def _phone_keys_in_cell(raw, strict: bool = False) -> list:
    """
    Todas as chaves (ddd, sub8) de telefone presentes numa célula — trata vários
    números separados por /, ; , etc. e números embutidos em texto.
    strict=True exige que cada token pareça telefone real (varredura de soltos).
    """
    s = str(raw)
    if not s or s.strip().lower() in ("nan", "none", ""):
        return []
    keys, seen = [], set()
    for tok in _PHONE_TOKEN_RE.findall(s):
        if strict and not _looks_like_phone(tok):
            continue
        k = phone_key(tok)
        if k[1] and k[1] not in seen:
            keys.append(k)
            seen.add(k[1])
    # Fallback célula inteira (ex.: formato que o regex não tokenizou). Em strict
    # só aceita se a célula inteira parece telefone real — evita pegar lixo.
    if not keys and (not strict or _looks_like_phone(s)):
        k = phone_key(s)
        if k[1]:
            keys.append(k)
    return keys


def _kommo_row_phone_keys(df: pd.DataFrame, phone_col: str) -> list:
    """Chaves (ddd, sub8) de telefone por linha do Kommo, coluna principal primeiro.

    Depois das colunas de telefone, RESGATA números escondidos nas demais colunas
    (ex.: telefone gravado no campo 'Email comercial' como 'p:+55...') — espelho
    da varredura de soltos que já existia no lado das vendas. Solto exige DDD.
    """
    phone_cols = [phone_col] + [
        c for c in df.columns
        if c != phone_col and not _is_doc_col(c) and not str(c).startswith("_")
        and _is_phone_col(df[c].dropna())
    ]
    row_keys: list = [[] for _ in range(len(df))]
    for col in phone_cols:
        if col not in df.columns:
            continue
        for i, raw in enumerate(df[col].tolist()):
            for key in _phone_keys_in_cell(raw):
                if all(key[1] != k[1] for k, _ in row_keys[i]):
                    row_keys[i].append((key, col))
    phone_set = set(phone_cols)
    for col in df.columns:
        if col in phone_set or _is_doc_col(col) or str(col).startswith("_"):
            continue
        for i, raw in enumerate(df[col].tolist()):
            s = str(raw)
            if len(s) < 10 or len(re.sub(r"\D", "", s)) < 10:
                continue
            for key in _phone_keys_in_cell(raw, strict=True):
                if key[0] is None:
                    continue  # telefone solto só entra com DDD (precisão)
                if all(key[1] != k[1] for k, _ in row_keys[i]):
                    row_keys[i].append((key, f"solto em {col}"))
    return row_keys


def _build_extended_lookup(
    df: pd.DataFrame, ds_with_meta: pd.DataFrame
) -> tuple[dict, dict, int]:
    """
    Retorna (lookup, lookup_source, n_stray):
      lookup        → {sub8: [(ddd, row_dict), ...]}  — DDD-aware; ordem = prioridade
      lookup_source → {sub8: nome_da_coluna_que_gerou_a_chave}
      n_stray       → nº de telefones soltos achados FORA das colunas de telefone
    Colunas de telefone primeiro (prioridade); depois varredura célula-a-célula de
    telefones soltos. Pula colunas de documento. Abas primárias têm prioridade (df já ordenado).
    """
    lookup: dict = {}
    lookup_source: dict = {}
    rows_dict = ds_with_meta.to_dict("index")  # pré-computa uma vez
    for _idx, _row in rows_dict.items():
        _row["_venda_idx"] = _idx  # identidade da linha de venda — dedup por VENDA

    phone_cols = []
    for col in df.columns:
        if _is_doc_col(col) or str(col).startswith("_"):
            continue
        col_data = df[col].dropna()
        if not _is_phone_col(col_data):
            continue
        phone_cols.append(col)
        for idx, raw in col_data.items():
            for key in _phone_keys_in_cell(raw):
                add_to_phone_lookup(lookup, key, rows_dict[idx])
                lookup_source.setdefault(key[1], col)

    # ── Varredura de telefones soltos (número fora da coluna de telefone) ──
    n_stray = 0
    phone_set = set(phone_cols)
    for col in df.columns:
        if col in phone_set or _is_doc_col(col) or str(col).startswith("_"):
            continue
        col_data = df[col].dropna()
        for idx, raw in col_data.items():
            if len(re.sub(r"\D", "", str(raw))) < 10:
                continue  # solto exige DDD (10-11 díg) — pré-filtro barato
            for ddd, sub8 in _phone_keys_in_cell(raw, strict=True):
                if ddd is None:
                    continue
                if sub8 in lookup and any(d == ddd for d, _ in lookup[sub8]):
                    continue  # já coberto por uma coluna de telefone
                add_to_phone_lookup(lookup, (ddd, sub8), rows_dict[idx])
                lookup_source.setdefault(sub8, f"solto em {col}")
                n_stray += 1
    return lookup, lookup_source, n_stray


def run_procv(
    df_sales: pd.DataFrame,
    sales_phone_col: str,
    df_kommo: pd.DataFrame,
    kommo_phone_col: str,
    kommo_tag_col: str,
    traffic_keyword: str,
    sales_name_col: Optional[str] = None,
    kommo_name_col: Optional[str] = None,
    traffic_exclude: str = "",
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """
    Retorna: (vendas_tratada, kommo_tratada, resultado_trafego, resultado_completo)
    Hierarquia de match: 1) Telefone principal  2) Telefone alternativo  3) Nome completo
    """
    ds = df_sales.copy()
    pos = ds.columns.get_loc(sales_phone_col) + 1
    ds.insert(pos, "Tel_Limpo_Vendas", ds[sales_phone_col].apply(clean_phone))
    ds.insert(pos + 1, "Tel_8dig_Vendas", ds["Tel_Limpo_Vendas"].apply(right8))

    dk = df_kommo.copy()
    pos_k = dk.columns.get_loc(kommo_phone_col) + 1
    dk.insert(pos_k, "Tel_Limpo_Kommo", dk[kommo_phone_col].apply(clean_phone))
    dk.insert(pos_k + 1, "Tel_8dig_Kommo", dk["Tel_Limpo_Kommo"].apply(right8))

    extended_lookup, lookup_source, n_stray_sales = _build_extended_lookup(df_sales, ds)

    # Lookup por nome: só nomes únicos nas vendas (nomes ambíguos são descartados)
    # word_index: palavra → conjunto de nomes normalizados — permite match parcial eficiente
    name_lookup: dict = {}
    name_word_index: dict = {}
    if sales_name_col and sales_name_col in df_sales.columns:
        _rows_dict = ds.to_dict("index")
        for _i2, _r2 in _rows_dict.items():
            _r2["_venda_idx"] = _i2
        _name_count: dict = {}
        for idx in df_sales.index:
            raw = df_sales.at[idx, sales_name_col]
            if pd.isna(raw):
                continue
            norm = _normalize_name(str(raw))
            # Nome de 1 palavra ("MARCIA") é ambíguo demais — casaria com
            # qualquer homônimo entre milhares de leads. Precisão > recall.
            if norm and len(norm.split()) >= 2:
                _name_count[norm] = _name_count.get(norm, 0) + 1
                if _name_count[norm] == 1:
                    name_lookup[norm] = _rows_dict[idx]
                else:
                    name_lookup.pop(norm, None)  # ambíguo — remove
        for norm in name_lookup:
            for word in norm.split():
                name_word_index.setdefault(word, set()).add(norm)

    # Pré-computa chaves de telefone (ddd, sub8) das colunas de telefone do Kommo por linha.
    # Lista de (key, col_name) por linha — coluna principal primeiro (prioridade).
    kommo_row_keys = _kommo_row_phone_keys(dk, kommo_phone_col)

    # Pré-extrai colunas como listas — iterar listas é MUITO mais rápido que iterrows()
    _k8_col = dk["Tel_8dig_Kommo"].tolist()
    _tag_col = dk[kommo_tag_col].tolist() if kommo_tag_col in dk.columns else [None] * len(dk)
    _phone_col = dk[kommo_phone_col].tolist() if kommo_phone_col in dk.columns else [""] * len(dk)
    _name_col = (dk[kommo_name_col].tolist()
                 if (kommo_name_col and kommo_name_col in dk.columns) else None)
    sales_cols = list(df_sales.columns)

    result_rows = []
    n_ddd_blocked = 0  # falsos matches evitados (sub8 igual mas DDD diferente)
    for i in range(len(dk)):
        k8_primary = _k8_col[i] or ""
        tag_raw = "" if pd.isna(_tag_col[i]) else str(_tag_col[i])
        phone_raw = _phone_col[i]
        is_traffic = _tag_matches(tag_raw, traffic_keyword, traffic_exclude)

        # Match por telefone DDD-aware: primeira chave (col principal → alternativas) que casa
        match_reason = ""
        sales_match = None
        matched_sub8 = ""
        matched_ddd = None
        blocked_here = False
        for key, kommo_col in kommo_row_keys[i]:
            cand = resolve_phone(extended_lookup, key)
            if not cand:
                if key[1] in extended_lookup:
                    blocked_here = True  # mesmo final, mas DDD não bate
                continue
            sales_match = cand
            matched_sub8 = key[1]
            matched_ddd = key[0]
            sales_col = lookup_source.get(key[1], "")
            if str(sales_col).startswith("solto em "):
                match_reason = f"Telefone solto nas vendas ({sales_col[9:]}) ⚠️ verificar"
            elif str(kommo_col).startswith("solto em "):
                match_reason = f"Telefone fora da coluna no Kommo ({str(kommo_col)[9:]}) ⚠️ verificar"
            elif kommo_col == kommo_phone_col and sales_col == sales_phone_col:
                match_reason = "Telefone"
            elif kommo_col == kommo_phone_col:
                match_reason = f"Telefone (vendas: {sales_col})"
            else:
                match_reason = f"Col. alt. Kommo: {kommo_col} → vendas: {sales_col}"
            break

        if not sales_match and _name_col is not None and name_lookup:
            kommo_name_raw = str(_name_col[i])
            kommo_norm = _normalize_name(kommo_name_raw)
            if kommo_norm:
                if kommo_norm in name_lookup and _nome_distintivo(kommo_norm.split()):
                    # Exact normalized match (com ao menos uma palavra distintiva)
                    sales_match = name_lookup[kommo_norm]
                    match_reason = "Nome completo ⚠️ verificar"
                else:
                    # Partial word match: intersecta candidatos por palavra do nome
                    kommo_words = kommo_norm.split()
                    if len(kommo_words) >= 2:
                        candidates: Optional[set] = None
                        for _w in kommo_words:
                            hits = name_word_index.get(_w)
                            if hits:
                                candidates = hits.copy() if candidates is None else candidates & hits
                        if candidates and len(candidates) == 1:
                            cand = next(iter(candidates))
                            if _names_match(kommo_norm, cand):
                                sales_match = name_lookup[cand]
                                match_reason = "Nome parcial ⚠️ verificar"

        if sales_match is None and blocked_here:
            n_ddd_blocked += 1

        row_out = {
            "Tag_Kommo": tag_raw,
            "Telefone_Kommo": phone_raw,
            "Nome_Kommo": ("" if (_name_col is None or pd.isna(_name_col[i]))
                           else str(_name_col[i])),
            "Tel_8dig": _row_ident(matched_ddd, matched_sub8, phone_raw),
            "É_Tráfego": "SIM" if is_traffic else "NÃO",
            "Venda_Confirmada": "SIM" if sales_match else "NÃO",
            "Criterio_Match": match_reason,
            "_Venda_Idx": str(sales_match.get("_venda_idx", "")) if sales_match else "",
        }
        if sales_match:
            for col in sales_cols:
                row_out[f"[Venda] {col}"] = sales_match.get(col, "")
        else:
            for col in sales_cols:
                row_out[f"[Venda] {col}"] = ""

        result_rows.append(row_out)

    df_full = pd.DataFrame(result_rows)
    if n_stray_sales:
        df_full.attrs["n_stray_sales"] = n_stray_sales
    if n_ddd_blocked:
        df_full.attrs["n_ddd_blocked"] = n_ddd_blocked
    df_trafego = df_full[
        (df_full["É_Tráfego"] == "SIM") & (df_full["Venda_Confirmada"] == "SIM")
    ].copy()
    # Deduplica por comprador único — mesmo telefone em vários leads Kommo não conta N vezes.
    # Leads sem Tel_8dig válido não participam do dedup (senão todos "" colapsam num só).
    if len(df_trafego) > 0:
        mask_tel = df_trafego["Tel_8dig"].str.len() > 0
        df_com_tel = df_trafego[mask_tel].drop_duplicates(subset=["Tel_8dig"])
        df_sem_tel = df_trafego[~mask_tel]
        df_trafego = pd.concat([df_com_tel, df_sem_tel], ignore_index=True)
        # A MESMA venda casada por números diferentes (fixo numa coluna, celular
        # noutra — mesma pessoa) também conta uma vez só.
        mask_idx = df_trafego["_Venda_Idx"].str.len() > 0
        df_trafego = pd.concat([
            df_trafego[mask_idx].drop_duplicates(subset=["_Venda_Idx"]),
            df_trafego[~mask_idx],
        ], ignore_index=True)

    # Coluna interna de dedup não vai para o relatório (preserva attrs no drop)
    df_trafego = df_trafego.drop(columns=["_Venda_Idx"], errors="ignore")
    _full_attrs = dict(df_full.attrs)
    df_full = df_full.drop(columns=["_Venda_Idx"], errors="ignore")
    df_full.attrs.update(_full_attrs)

    return ds, dk, df_trafego, df_full


# ── Utilitários de data ────────────────────────────────────────────────────────

_MONTH_MAP = {
    "janeiro": 1, "jan": 1, "fevereiro": 2, "fev": 2,
    "março": 3, "marco": 3, "mar": 3, "abril": 4, "abr": 4,
    "maio": 5, "mai": 5, "junho": 6, "jun": 6,
    "julho": 7, "jul": 7, "agosto": 8, "ago": 8,
    "setembro": 9, "set": 9, "outubro": 10, "out": 10,
    "novembro": 11, "nov": 11, "dezembro": 12, "dez": 12,
    "january": 1, "february": 2, "feb": 2, "march": 3, "april": 4, "apr": 4,
    "may": 5, "june": 6, "july": 7, "august": 8, "aug": 8,
    "september": 9, "sep": 9, "october": 10, "oct": 10,
    "november": 11, "december": 12, "dec": 12,
}

def _normalize(s: str) -> str:
    return (s.lower()
            .replace("ç", "c").replace("ã", "a").replace("á", "a")
            .replace("â", "a").replace("à", "a").replace("é", "e").replace("ê", "e")
            .replace("í", "i").replace("ó", "o").replace("ô", "o")
            .replace("ú", "u").replace("ü", "u").replace("õ", "o"))


def _tag_matches(tag_text, include, exclude="") -> bool:
    """True se a tag bate com QUALQUER palavra de `include` (separadas por , ou ;)
    E não bate com NENHUMA palavra de `exclude`. Ignora acentos e maiúsculas.

    Ex.: include='trafego, pago, ads' casa qualquer uma das três;
         include='trafego', exclude='organico' casa 'Tráfego Pago' mas não 'Tráfego Orgânico'.
    """
    norm = _normalize(str(tag_text))
    incs = [k.strip() for k in re.split(r"[;,]", _normalize(str(include))) if k.strip()]
    if not incs or not any(k in norm for k in incs):
        return False
    excs = [k.strip() for k in re.split(r"[;,]", _normalize(str(exclude))) if k.strip()]
    return not any(k in norm for k in excs)


_NAME_STOP = {"de", "da", "do", "dos", "das", "e", "di", "du", "van", "von", "el", "la"}

def _normalize_name(name: str) -> str:
    """Minúsculas, sem acentos, sem artigos/preposições. Ex: 'Dirce Natalina de Vesselai' → 'dirce natalina vesselai'."""
    s = _normalize(str(name).strip())
    return " ".join(p for p in s.split() if p not in _NAME_STOP and len(p) > 1)


# Nomes/sobrenomes mais comuns do Brasil: um match composto SÓ por eles não
# identifica ninguém — "Maria José" × "Maria José" casa milhares de homônimos.
_NOMES_GENERICOS = {
    "maria", "jose", "ana", "joao", "antonio", "francisco", "carlos", "paulo",
    "pedro", "lucas", "luiz", "luis", "marcos", "gabriel", "rafael", "daniel",
    "marcelo", "bruno", "eduardo", "felipe", "fernanda", "juliana", "adriana",
    "patricia", "aline", "camila", "amanda", "bruna", "leticia", "sandra",
    "aparecida", "fatima", "lucia", "vera", "paula", "claudia",
    "silva", "santos", "souza", "sousa", "oliveira", "costa", "pereira",
    "lima", "alves", "ferreira", "rodrigues", "almeida", "nascimento",
    "carvalho", "araujo", "ribeiro", "gomes", "martins", "barbosa", "rocha",
    "dias", "nunes", "moreira", "cavalcante", "cavalcanti", "correia",
    "junior", "filho", "neto",
}


def _nome_distintivo(tokens) -> bool:
    """True se há ao menos UMA palavra fora da lista de nomes ultra-comuns."""
    return any(t not in _NOMES_GENERICOS for t in tokens)


def _names_match(name_a: str, name_b: str) -> bool:
    """
    True se os nomes normalizados são equivalentes.
    Exige ≥ 2 palavras em comum cobrindo todo o nome mais curto, e que ao menos
    uma palavra em comum seja DISTINTIVA (fora dos nomes mais comuns do Brasil) —
    'Maria José' × 'Maria José Fonseca' não casa; 'Leila Puzzi' × 'Leila Puzzi da Silva' casa.
    Nomes de 1 palavra são ambíguos demais — não são considerados match.
    """
    na = _normalize_name(name_a)
    nb = _normalize_name(name_b)
    if not na or not nb:
        return False
    pa, pb = set(na.split()), set(nb.split())
    if len(pa) < 2 or len(pb) < 2:
        return False
    common = pa & pb
    if len(common) < min(len(pa), len(pb)) or len(common) < 2:
        return False
    return _nome_distintivo(common)


def _is_real_datetime(v) -> bool:
    """Retorna True apenas para datetime real — exclui pd.NaT que herda de datetime."""
    if v is None or v is pd.NaT:
        return False
    try:
        v.strftime("%Y")
        return True
    except Exception:
        return False


def parse_date(value) -> Optional[datetime]:
    if value is None or value is pd.NaT:
        return None
    if isinstance(value, datetime) and _is_real_datetime(value):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime() if _is_real_datetime(value) else None

    s = str(value).strip()
    if not s or s.lower() in ("nan", "none", "-", "n/a", "#n/a", ""):
        return None

    # Excel serial number como string (ex: "45678")
    try:
        serial = int(float(s))
        if 35000 < serial < 60000:
            return datetime(1899, 12, 30) + timedelta(days=serial)
    except (ValueError, OverflowError):
        pass

    # Formato DD/MM/YY ou DD/MM/YYYY (padrão BR) — fallback MM/DD se dia inválido
    # Ex: "22/05/26" → May 22; "4/30/26" → April 30 (dia 4 mês 30 inválido, inverte)
    m_us = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{2,4})(?:\s+\d{1,2}:\d{2})?$", s)
    if m_us:
        a, b, yr = int(m_us.group(1)), int(m_us.group(2)), int(m_us.group(3))
        if yr < 100:
            yr += 2000
        # Tenta DD/MM/YYYY primeiro (padrão brasileiro)
        try:
            if 1 <= a <= 31 and 1 <= b <= 12:
                return datetime(yr, b, a)
        except ValueError:
            pass
        # Fallback MM/DD/YYYY (ex: "4/30/26" onde dia=30 > 12)
        try:
            if 1 <= b <= 31 and 1 <= a <= 12:
                return datetime(yr, a, b)
        except ValueError:
            pass

    # Formatos padrão via pandas — tenta dayfirst=False primeiro para M/D/YY
    for dayfirst in (False, True):
        try:
            dt = pd.to_datetime(s, dayfirst=dayfirst)
            if _is_real_datetime(dt):
                return dt.to_pydatetime()
        except Exception:
            continue

    # "month_name/YY", "month_name/YYYY", "month_name YYYY" (ex: "abril/26", "março 2024")
    sn = _normalize(s)
    m = re.match(r"^([a-z]+)[/\s\-](\d{2,4})$", sn)
    if m:
        month = _MONTH_MAP.get(m.group(1))
        if month:
            year = int(m.group(2))
            if year < 100:
                year += 2000
            try:
                return datetime(year, month, 1)
            except ValueError:
                pass

    # "DD de month_name de YYYY" / "DD month_name YYYY" / "22-abr-2026" / "22/abril/26"
    m = re.match(r"^(\d{1,2})[\s/\-]+(?:de[\s/\-]+)?([a-z]+)(?:[\s/\-]+(?:de[\s/\-]+)?(\d{2,4}))?$", sn)
    if m:
        month = _MONTH_MAP.get(m.group(2))
        if month:
            year = int(m.group(3)) if m.group(3) else datetime.now().year
            if year < 100:
                year += 2000
            try:
                return datetime(year, month, int(m.group(1)))
            except ValueError:
                pass

    # Extrai data embutida em texto livre (ex: "Disparo dia das Mães 22/04" ou "23/04/2026 17:42")
    dm = re.search(r"\b(\d{1,2})[/\-\.](\d{1,2})(?:[/\-\.](\d{2,4}))?\b", s)
    if dm:
        try:
            day, month = int(dm.group(1)), int(dm.group(2))
            if 1 <= day <= 31 and 1 <= month <= 12:
                if dm.group(3):
                    year = int(dm.group(3))
                    if year < 100:
                        year += 2000
                else:
                    # Sem ano: infere pelo mês atual
                    # Se o mês extraído > mês atual, é do ano anterior
                    now = datetime.now()
                    year = now.year if month <= now.month else now.year - 1
                return datetime(year, month, day)
        except (ValueError, TypeError):
            pass

    return None


@st.cache_data(show_spinner=False)
def detect_date_col(df: pd.DataFrame) -> Optional[str]:
    """
    Detecta coluna de data por conteúdo primeiro, nome como bônus.
    Funciona com colunas chamadas "Data", "Dt", "Date", "Período", etc.
    Cada planilha é um caso — não assume padronização.
    """
    kws = [
        "data", "date", "dt", "dia", "quando", "periodo", "período",
        "competencia", "competência", "venda_em", "criado", "created",
        "hora", "timestamp", "tempo", "time", "mes", "mês", "ano",
        "year", "month", "compra", "pedido", "emissao", "emissão",
        "lancamento", "lançamento", "movimento", "referencia", "referência",
    ]
    # Colunas que provavelmente NÃO são datas (evita falsos positivos)
    skip_kws = ["telefone", "celular", "phone", "tel", "whatsapp", "wpp",
                "valor", "preco", "preço", "total", "cpf", "cnpj", "cep"]

    # Valor monetário ("76915.55", "1.234,56") não é data — o parser de serial
    # Excel transformaria dinheiro na faixa 35000-60000 em datas fantasma.
    _money_re = re.compile(r"R?\$?\s*-?(?:\d{1,3}(?:\.\d{3})*,\d{1,2}|\d+\.\d+)")

    candidates: list = []
    for col in df.columns:
        if str(col).startswith("_") or _match_keywords(col, skip_kws):
            continue
        col_nonnull = df[col].dropna()
        # Meia dúzia de células soltas (data de emissão no rodapé do relatório)
        # não é a coluna de data das vendas.
        if len(df) > 30 and len(col_nonnull) < 3:
            continue
        sample = col_nonnull.head(25)
        if len(sample) == 0:
            continue
        money_like = sum(1 for v in sample if _money_re.fullmatch(str(v).strip()))
        if money_like / len(sample) >= 0.4:
            continue
        hits = sum(1 for v in sample if parse_date(v) is not None)
        if hits < 2:
            continue
        content_score = int((hits / len(sample)) * 60)  # 0-60
        name_bonus = 40 if _match_keywords(col, kws) else 0
        # Coluna constante (ex.: "período: 01/06") não é a data da venda
        variance_penalty = 30 if (len(sample) >= 5 and sample.astype(str).nunique() == 1) else 0
        candidates.append((col, content_score + name_bonus - variance_penalty))

    if not candidates:
        return None
    # Retorna a coluna com maior pontuação combinada (conteúdo + nome)
    return max(candidates, key=lambda x: x[1])[0]


# ── Análise de Disparo ─────────────────────────────────────────────────────────

def run_disparo(
    df_sales: pd.DataFrame,
    sales_phone_col: str,
    sales_date_col: Optional[str],
    df_kommo: pd.DataFrame,
    kommo_phone_col: str,
    kommo_tag_col: str,
    disparo_keyword: str,
    kommo_date_col: Optional[str],
    window_days: int = 30,
    sales_name_col: Optional[str] = None,
    kommo_name_col: Optional[str] = None,
    disparo_exclude: str = "",
) -> pd.DataFrame:
    """
    Filtra leads de disparo do Kommo pela tag e cruza com vendas.
    Janela de tempo: venda entre 0 e window_days após a data do disparo no Kommo.
    Hierarquia de match: 1) Telefone  2) Telefone alternativo  3) Nome completo/parcial
    """
    # Filtra leads de disparo no Kommo (ignora acentos e maiúsculas)
    disparo_mask = df_kommo[kommo_tag_col].fillna("").astype(str).apply(
        lambda t: _tag_matches(t, disparo_keyword, disparo_exclude)
    )
    df_disp_leads = df_kommo[disparo_mask].copy()

    if len(df_disp_leads) == 0:
        return pd.DataFrame()

    ds = df_sales.copy()
    if sales_date_col:
        ds["_dt_venda"] = ds[sales_date_col].apply(parse_date)

    # Lookup estendido DDD-aware: {sub8: [(ddd, sale_dict), ...]} — todas as colunas de
    # telefone das vendas + telefones soltos. added_sale_keys evita duplicar a mesma linha.
    added_sale_keys: set = set()    # (sub8, idx)
    sales_lookup: dict = {}
    sales_lookup_source: dict = {}  # {sub8: col_name}

    def _sale_payload(idx):
        d = ds.loc[idx].to_dict()
        d["_venda_idx"] = idx  # identidade da linha de venda — dedup por VENDA
        return d
    phone_cols_sales = [c for c in df_sales.columns
                        if not _is_doc_col(c) and _is_phone_col(df_sales[c].dropna())]
    for col in phone_cols_sales:
        for idx, raw in df_sales[col].dropna().items():
            for ddd, sub8 in _phone_keys_in_cell(raw):
                if (sub8, idx) in added_sale_keys:
                    continue
                sales_lookup.setdefault(sub8, []).append((ddd, _sale_payload(idx)))
                added_sale_keys.add((sub8, idx))
                sales_lookup_source.setdefault(sub8, col)
    # Telefones soltos fora das colunas de telefone (número fora da coluna)
    for col in df_sales.columns:
        if col in phone_cols_sales or _is_doc_col(col):
            continue
        for idx, raw in df_sales[col].dropna().items():
            if len(re.sub(r"\D", "", str(raw))) < 10:
                continue
            for ddd, sub8 in _phone_keys_in_cell(raw, strict=True):
                if ddd is None or (sub8, idx) in added_sale_keys:
                    continue
                sales_lookup.setdefault(sub8, []).append((ddd, _sale_payload(idx)))
                added_sale_keys.add((sub8, idx))
                sales_lookup_source.setdefault(sub8, f"solto em {col}")

    # Lookup por nome para fallback (nomes únicos nas vendas)
    disp_name_lookup: dict = {}
    disp_name_word_index: dict = {}
    if sales_name_col and sales_name_col in df_sales.columns:
        _ds_rows = ds.to_dict("index")
        for _i3, _r3 in _ds_rows.items():
            _r3["_venda_idx"] = _i3
        _nc: dict = {}
        for _idx in df_sales.index:
            _raw = df_sales.at[_idx, sales_name_col]
            if pd.isna(_raw):
                continue
            _norm = _normalize_name(str(_raw))
            # 1 palavra é ambíguo demais (ver run_procv) — precisão > recall
            if _norm and len(_norm.split()) >= 2:
                _nc[_norm] = _nc.get(_norm, 0) + 1
                if _nc[_norm] == 1:
                    disp_name_lookup[_norm] = _ds_rows[_idx]
                else:
                    disp_name_lookup.pop(_norm, None)
        for _n in disp_name_lookup:
            for _w in _n.split():
                disp_name_word_index.setdefault(_w, set()).add(_n)

    # Pré-computa chaves (ddd, sub8) das colunas de telefone para cada lead de disparo.
    # Lista de (key, col_name) por linha — coluna principal primeiro (prioridade).
    df_disp_reset = df_disp_leads.reset_index(drop=True)
    disp_row_keys = _kommo_row_phone_keys(df_disp_reset, kommo_phone_col)

    # Pré-extrai colunas como listas (evita iterrows()) — chave pra escalar no breakdown
    _d_phone = df_disp_reset[kommo_phone_col].tolist() if kommo_phone_col in df_disp_reset.columns else [""] * len(df_disp_reset)
    _d_tag = df_disp_reset[kommo_tag_col].tolist() if kommo_tag_col in df_disp_reset.columns else [""] * len(df_disp_reset)
    _d_date = (df_disp_reset[kommo_date_col].tolist()
               if (kommo_date_col and kommo_date_col in df_disp_reset.columns) else None)
    _d_name = (df_disp_reset[kommo_name_col].tolist()
               if (kommo_name_col and kommo_name_col in df_disp_reset.columns) else None)
    sales_cols_d = list(df_sales.columns)

    result_rows = []
    for _i in range(len(df_disp_reset)):
        tel_raw = _d_phone[_i]

        row_keys = disp_row_keys[_i]
        has_phone = bool(row_keys)
        has_name_fallback = bool(kommo_name_col and disp_name_lookup)
        if not has_phone and not has_name_fallback:
            continue

        # Data do disparo: 1) data no texto da PRÓPRIA tag de disparo (ex.:
        # "DISPARO NAMORADOS 09/06/26" — sinal direto da campanha), 2) coluna de
        # data do Kommo (geralmente "Criado em", que pode ser bem anterior ao
        # disparo em leads importados), 3) qualquer data no texto das tags.
        _tag_txt = "" if pd.isna(_d_tag[_i]) else str(_d_tag[_i])
        disp_date = None
        for _seg in _tag_txt.split(","):
            if _tag_matches(_seg, disparo_keyword, disparo_exclude):
                _sd = parse_date(_seg)
                if _is_real_datetime(_sd):
                    disp_date = _sd
                    break
        if not _is_real_datetime(disp_date) and _d_date is not None:
            disp_date = parse_date(_d_date[_i])
        if not _is_real_datetime(disp_date):
            disp_date = parse_date(_tag_txt)

        confirmed_sale = None
        phone_only_sale = None
        match_reason = ""
        matched_sub8 = ""
        matched_ddd = None
        any_phone_sale = False  # o telefone do lead bate com alguma venda?

        # ── Match por telefone DDD-aware ──────────────────────────────────
        for key, k_col in row_keys:
            if confirmed_sale:
                break
            sub8 = key[1]
            s_col = sales_lookup_source.get(sub8, sales_phone_col)
            if str(s_col).startswith("solto em "):
                _mr = f"Telefone solto nas vendas ({s_col[9:]}) ⚠️ verificar"
            elif str(k_col).startswith("solto em "):
                _mr = f"Telefone fora da coluna no Kommo ({str(k_col)[9:]}) ⚠️ verificar"
            elif k_col == kommo_phone_col and s_col == sales_phone_col:
                _mr = "Telefone"
            elif k_col == kommo_phone_col:
                _mr = f"Telefone (vendas: {s_col})"
            else:
                _mr = f"Col. alt. Kommo: {k_col} → vendas: {s_col}"
            sales_for_key = resolve_phone_all(sales_lookup, key)
            if sales_for_key:
                any_phone_sale = True
            for sale in sales_for_key:
                if _is_real_datetime(disp_date) and sales_date_col:
                    sale_dt = sale.get("_dt_venda")
                    if _is_real_datetime(sale_dt):
                        delta = (sale_dt - disp_date).days
                        if 0 <= delta <= window_days:
                            confirmed_sale = sale
                            match_reason = _mr
                            matched_sub8 = sub8; matched_ddd = key[0]
                            break
                    elif phone_only_sale is None:
                        phone_only_sale = sale
                        match_reason = _mr
                        matched_sub8 = sub8; matched_ddd = key[0]
                else:
                    confirmed_sale = sale
                    match_reason = _mr
                    matched_sub8 = sub8; matched_ddd = key[0]
                    break

        # ── Fallback: match por nome (só quando telefone falhou) ─────────
        if not confirmed_sale and not phone_only_sale and has_name_fallback:
            kommo_name_raw = str(_d_name[_i]) if _d_name is not None else ""
            kommo_norm = _normalize_name(kommo_name_raw)
            name_candidate = None
            _nr = ""
            if kommo_norm:
                if kommo_norm in disp_name_lookup and _nome_distintivo(kommo_norm.split()):
                    name_candidate = disp_name_lookup[kommo_norm]
                    _nr = "Nome completo ⚠️ verificar"
                else:
                    _dw = kommo_norm.split()
                    if len(_dw) >= 2:
                        _nc2: Optional[set] = None
                        for _w in _dw:
                            _h = disp_name_word_index.get(_w)
                            if _h:
                                _nc2 = _h.copy() if _nc2 is None else _nc2 & _h
                        if _nc2 and len(_nc2) == 1:
                            _cand = next(iter(_nc2))
                            if _names_match(kommo_norm, _cand):
                                name_candidate = disp_name_lookup[_cand]
                                _nr = "Nome parcial ⚠️ verificar"
            if name_candidate:
                if _is_real_datetime(disp_date) and sales_date_col:
                    sale_dt = name_candidate.get("_dt_venda")
                    if _is_real_datetime(sale_dt):
                        delta = (sale_dt - disp_date).days
                        if 0 <= delta <= window_days:
                            confirmed_sale = name_candidate
                            match_reason = _nr
                    else:
                        phone_only_sale = name_candidate
                        match_reason = _nr
                else:
                    confirmed_sale = name_candidate
                    match_reason = _nr

        matched_sale = confirmed_sale if confirmed_sale else phone_only_sale

        tag_raw = str(_d_tag[_i]) if not pd.isna(_d_tag[_i]) else ""
        row_out: dict = {
            "Tag_Kommo": tag_raw,
            "Telefone_Disparo": tel_raw,
            "Nome_Kommo": ("" if (_d_name is None or pd.isna(_d_name[_i]))
                           else str(_d_name[_i])),
            "Tel_8dig": _row_ident(matched_ddd, matched_sub8, tel_raw),
            "_Venda_Idx": str(matched_sale.get("_venda_idx", "")) if matched_sale else "",
        }

        if kommo_date_col:
            row_out["Data_Disparo"] = _d_date[_i] if _d_date is not None else ""

        if matched_sale:
            row_out["Venda_Confirmada"] = "SIM"
            row_out["Criterio_Match"] = match_reason
            if sales_date_col:
                sale_dt = matched_sale.get("_dt_venda")
                if _is_real_datetime(sale_dt):
                    row_out["Data_Venda"] = sale_dt.strftime("%d/%m/%Y")
                    if _is_real_datetime(disp_date):
                        row_out["Dias_Após_Disparo"] = (sale_dt - disp_date).days
                else:
                    row_out["Data_Venda"] = str(matched_sale.get(sales_date_col, ""))
            for col in sales_cols_d:
                row_out[f"[Venda] {col}"] = matched_sale.get(col, "")
        else:
            row_out["Venda_Confirmada"] = "NÃO"
            # Motivo da não-conversão — ajuda a entender o resultado sem adivinhar
            if any_phone_sale:
                row_out["Criterio_Match"] = "Telefone bate, mas a venda ficou fora da janela (antes do disparo ou +30 dias)"
            else:
                row_out["Criterio_Match"] = "Telefone do lead não encontrado nas vendas"
            if sales_date_col:
                row_out["Data_Venda"] = ""
                if disp_date:
                    row_out["Dias_Após_Disparo"] = ""
            for col in sales_cols_d:
                row_out[f"[Venda] {col}"] = ""

        result_rows.append(row_out)

    if not result_rows:
        return pd.DataFrame()
    df_res = pd.DataFrame(result_rows)
    # Deduplica por Tel_8dig: mesmo comprador em múltiplos leads Kommo não conta N vezes.
    # Prioriza Venda_Confirmada=SIM sobre NÃO antes de desduplicar.
    if "Tel_8dig" in df_res.columns:
        mask_tel = df_res["Tel_8dig"].str.len() > 0
        df_com_tel = (df_res[mask_tel]
                      .sort_values("Venda_Confirmada", ascending=False)  # SIM primeiro
                      .drop_duplicates(subset=["Tel_8dig"]))
        df_sem_tel = df_res[~mask_tel]
        df_res = pd.concat([df_com_tel, df_sem_tel], ignore_index=True)
        # Mesma venda casada por números diferentes da mesma pessoa → conta uma vez
        mask_idx = df_res["_Venda_Idx"].str.len() > 0
        df_res = pd.concat([
            df_res[mask_idx].drop_duplicates(subset=["_Venda_Idx"]),
            df_res[~mask_idx],
        ], ignore_index=True)
    return df_res.drop(columns=["_Venda_Idx"], errors="ignore")


# ── Breakdown por aba/mês ───────────────────────────────────────────────────────

def run_breakdown_by_sheet(
    df_sales: pd.DataFrame,
    sales_phone_col: str,
    df_kommo: pd.DataFrame,
    kommo_phone_col: str,
    kommo_tag_col: str,
    traffic_keyword: str,
    *,
    sales_date_col: Optional[str] = None,
    disparo_keyword: Optional[str] = None,
    kommo_date_col: Optional[str] = None,
    sales_name_col: Optional[str] = None,
    kommo_name_col: Optional[str] = None,
    df_kommo_disp: Optional[pd.DataFrame] = None,
    kommo_disp_phone_col: Optional[str] = None,
    kommo_disp_tag_col: Optional[str] = None,
    traffic_exclude: str = "",
    disparo_exclude: str = "",
) -> Optional[pd.DataFrame]:
    """
    Cruza UM Kommo contra CADA aba (mês) da planilha de vendas, separadamente.
    É o cenário "planilha de 8 meses × 1 Kommo → vendas do tráfego mês a mês".

    A deduplicação acontece DENTRO de cada mês — um comprador recorrente conta
    em cada mês em que comprou (ao contrário do total combinado, que o conta 1×).

    Retorna um DataFrame (uma linha por aba) ou None se a planilha tem 1 aba só.
    """
    if "_Planilha" not in df_sales.columns:
        return None

    rows = []
    for sheet, grp in df_sales.groupby("_Planilha", sort=False):
        g = grp.reset_index(drop=True)
        _, _, traf, full = run_procv(
            g, sales_phone_col, df_kommo, kommo_phone_col, kommo_tag_col, traffic_keyword,
            sales_name_col=sales_name_col, kommo_name_col=kommo_name_col,
            traffic_exclude=traffic_exclude,
        )
        row = {
            "Mês / Aba": sheet,
            "Vendas no mês": len(g),
            "Vendas de Tráfego": len(traf),
        }
        if disparo_keyword and disparo_keyword.strip():
            kd = df_kommo_disp if df_kommo_disp is not None else df_kommo
            kp = kommo_disp_phone_col or kommo_phone_col
            kt = kommo_disp_tag_col or kommo_tag_col
            disp = run_disparo(
                g, sales_phone_col, sales_date_col, kd, kp, kt, disparo_keyword, kommo_date_col,
                sales_name_col=sales_name_col, kommo_name_col=kommo_name_col,
                disparo_exclude=disparo_exclude,
            )
            row["Vendas de Disparo"] = int((disp["Venda_Confirmada"] == "SIM").sum()) if len(disp) else 0
        rows.append(row)

    return pd.DataFrame(rows) if rows else None


# ── Análise de Duplicatas ──────────────────────────────────────────────────────

def _is_id_col(col_data: pd.Series, col_name: str) -> bool:
    """Detecta colunas de ID sequencial que não devem entrar no fingerprint."""
    id_kws = ["id", "codigo", "código", "protocolo", "numero", "número",
              "registro", "pedido", "ordem", "seq", "nro", "nº", "num",
              "index", "índice", "indice", "linha", "row"]
    if any(kw in col_name.lower() for kw in id_kws):
        return True
    non_null = col_data.dropna()
    if len(non_null) < 4:
        return False
    if non_null.nunique() == len(non_null):
        is_int = non_null.apply(
            lambda v: str(v).strip().replace(".0", "").lstrip("0").isdigit()
        ).sum()
        if int(is_int) >= 0.8 * len(non_null):
            return True
    return False


def analyze_duplicates(
    df: pd.DataFrame,
    phone_col: str,
    date_col: Optional[str] = None,
) -> pd.DataFrame:
    """
    Classifica registros com telefone repetido.
    Adiciona coluna 'Situacao_Venda':
      Única               → só uma venda para este número
      Multi-compra        → mesmo número, datas diferentes (compras reais)
      Multi-compra (mesmo dia) → mesmo número, mesma data, conteúdo diferente
      Multi-compra (sem data)  → mesmo número, sem data, conteúdo diferente
      DUPLICATA           → mesmo número + mesma data + conteúdo idêntico
      DUPLICATA (sem data)→ mesmo número, sem data, conteúdo idêntico
    """
    result = df.copy()
    phones = df[phone_col].apply(
        lambda v: phone_group_key(v) if pd.notna(v) else ""
    )

    situacao = pd.Series(["Única"] * len(df), index=df.index, dtype=str)

    counts = phones[phones != ""].value_counts()
    multi_phones = counts[counts > 1].index

    id_cols = {col for col in df.columns if _is_id_col(df[col], col)}
    skip_set = {phone_col} | id_cols
    if date_col:
        skip_set.add(date_col)
    fp_cols = [c for c in df.columns if c not in skip_set]

    def _fingerprint(row) -> str:
        parts = []
        for c in fp_cols:
            v = row.get(c, "")
            if pd.isna(v):
                parts.append("")
                continue
            s = str(v).strip().lower()
            try:
                s = str(round(float(s.replace(",", ".").replace(" ", "")), 2))
            except ValueError:
                pass
            parts.append(s)
        return "|".join(parts)

    for phone in multi_phones:
        mask = phones == phone
        group = df[mask]

        if date_col and date_col in df.columns:
            def _fmt_date(v):
                d = parse_date(v)
                return d.strftime("%Y-%m-%d") if d else ""
            date_keys = group[date_col].apply(_fmt_date)
            unique_dates = set(d for d in date_keys if d)
        else:
            date_keys = pd.Series([""] * len(group), index=group.index)
            unique_dates = set()

        if len(unique_dates) > 1:
            situacao[mask] = "Multi-compra"
        elif len(unique_dates) == 1:
            fps = group.apply(_fingerprint, axis=1)
            if fps.nunique() == 1:
                situacao[mask] = "DUPLICATA"
            else:
                situacao[mask] = "Multi-compra (mesmo dia)"
        else:
            fps = group.apply(_fingerprint, axis=1)
            if fps.nunique() == 1:
                situacao[mask] = "DUPLICATA (sem data)"
            else:
                situacao[mask] = "Multi-compra (sem data)"

    result.insert(0, "Situacao_Venda", situacao)
    return result


# ── Geração do Excel formatado ─────────────────────────────────────────────────

def _col_width(df: pd.DataFrame, col: str) -> float:
    max_data = df[col].dropna().astype(str).apply(len).max() if len(df) > 0 else 10
    return min(float(max(len(str(col)), max_data)) + 4, 45)


def _write_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    header_hex: str,
    highlight_cols: Optional[List[str]] = None,
    max_rows: int = 10000,
):
    THIN = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Limita a LISTAGEM em bases enormes para o Excel não travar. Os totais (Resumo,
    # Mês a Mês, métricas na tela) são calculados sobre os dados COMPLETOS — só a
    # listagem detalhada aqui é recortada.
    n_full = len(df)
    if max_rows and n_full > max_rows:
        df = df.head(max_rows)
        title = f"{title}  —  primeiras {max_rows:,} de {n_full:,} linhas"

    # Título
    last_col = get_column_letter(max(len(df.columns), 1))
    ws.merge_cells(f"A1:{last_col}1")
    tc = ws.cell(1, 1, title)
    tc.font = Font(color="FFFFFF", bold=True, size=13)
    tc.fill = PatternFill("solid", fgColor=header_hex)
    tc.alignment = CENTER
    ws.row_dimensions[1].height = 32

    # Cabeçalhos
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(2, c, str(col))
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.fill = PatternFill("solid", fgColor="222222")
        cell.alignment = CENTER
        cell.border = THIN
    ws.row_dimensions[2].height = 24

    # Estilos pré-criados — evita instanciar N×M objetos dentro do loop
    _FONT_DATA      = Font(size=10)
    _FONT_WB        = Font(color="FFFFFF", bold=True, size=10)
    _FONT_DARK_B    = Font(color="1C1C1E", bold=True, size=10)
    _FILL_ZEBRA     = PatternFill("solid", fgColor="F5F5F5")
    _FILL_GREEN     = PatternFill("solid", fgColor="27AE60")
    _FILL_RED       = PatternFill("solid", fgColor="E74C3C")
    _FILL_BLUE_T    = PatternFill("solid", fgColor="2980B9")
    _FILL_BLUE_ALT  = PatternFill("solid", fgColor="5DADE2")
    _FILL_AMBER     = PatternFill("solid", fgColor="F39C12")
    _FILL_AMBER_ROW = PatternFill("solid", fgColor="FEF3E2")   # fundo de linha para revisão
    _FILL_AMBER_ROW_Z = PatternFill("solid", fgColor="FDEBD0") # versão zebra
    _FILL_DUP_RED   = PatternFill("solid", fgColor="C0392B")
    _FILL_DUP_ORA   = PatternFill("solid", fgColor="E67E22")
    _FILL_MULTI_B   = PatternFill("solid", fgColor="2471A3")
    _FILL_HL        = PatternFill("solid", fgColor="FFE5D9")
    _FILL_HL_Z      = PatternFill("solid", fgColor="FFD5B8")

    # Índice da coluna "Criterio_Match" para detectar match por nome no nível da linha
    _criterio_idx = list(df.columns).index("Criterio_Match") if "Criterio_Match" in df.columns else -1

    # Dados — só atribui estilo onde há COR de significado (verde/vermelho/zebra…).
    # Não põe borda/fonte/alinhamento por célula: a cada milhar de linhas isso era
    # o que travava a geração (eram milhões de objetos de estilo).
    highlight_set = set(highlight_cols or [])
    cols = list(df.columns)
    for r, row_data in enumerate(df.itertuples(index=False), 3):
        zebra = (r - 3) % 2 == 1
        row_vals = list(row_data)
        criterio_val = str(row_vals[_criterio_idx]) if _criterio_idx >= 0 else ""
        is_name_match = criterio_val.startswith("Nome")

        for c, value in enumerate(row_vals, 1):
            col_name = cols[c - 1]
            cell = ws.cell(r, c, value if value != "" else None)

            if col_name == "Venda_Confirmada":
                if value == "SIM":
                    # Match por nome: âmbar (precisa revisão). Match por tel: verde.
                    if is_name_match:
                        cell.fill = _FILL_AMBER
                        cell.font = _FONT_DARK_B
                    else:
                        cell.fill = _FILL_GREEN
                        cell.font = _FONT_WB
                elif value == "NÃO":
                    cell.fill = _FILL_RED
                    cell.font = _FONT_WB
            elif col_name == "Criterio_Match":
                sv = str(value)
                if sv == "Telefone":
                    cell.fill = _FILL_GREEN
                    cell.font = _FONT_WB
                elif sv == "Telefone (col. alternativa)":
                    cell.fill = _FILL_BLUE_ALT
                    cell.font = _FONT_WB
                elif sv.startswith("Nome"):
                    cell.fill = _FILL_AMBER
                    cell.font = _FONT_DARK_B
            elif col_name == "É_Tráfego" and value == "SIM":
                cell.fill = _FILL_BLUE_T
                cell.font = _FONT_WB
            elif col_name == "Situacao_Venda":
                sv = str(value)
                if sv.startswith("DUPLICATA"):
                    cell.fill = _FILL_DUP_RED
                    cell.font = _FONT_WB
                elif "mesmo dia" in sv or "sem data" in sv:
                    cell.fill = _FILL_DUP_ORA
                    cell.font = _FONT_WB
                elif sv.startswith("Multi"):
                    cell.fill = _FILL_MULTI_B
                    cell.font = _FONT_WB
                elif is_name_match:
                    cell.fill = _FILL_AMBER_ROW_Z if zebra else _FILL_AMBER_ROW
                elif zebra:
                    cell.fill = _FILL_ZEBRA
            elif col_name in highlight_set:
                cell.fill = _FILL_HL if not zebra else _FILL_HL_Z
            elif is_name_match:
                # Toda a linha com fundo âmbar claro para indicar revisão
                cell.fill = _FILL_AMBER_ROW_Z if zebra else _FILL_AMBER_ROW
            elif zebra:
                cell.fill = _FILL_ZEBRA

    # Largura das colunas
    for c, col in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(c)].width = _col_width(df, col)

    ws.freeze_panes = "A3"
    if len(df) > 0:
        ws.auto_filter.ref = f"A2:{last_col}{len(df)+2}"


def _rename_result_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Renomeia colunas técnicas para linguagem legível no Excel."""
    return df.rename(columns={
        "Venda_Confirmada": "Comprou?",
        "É_Tráfego": "É Lead de Tráfego?",
        "Tag_Kommo": "Tag no Kommo",
        "Telefone_Kommo": "Telefone (Kommo)",
        "Telefone_Disparo": "Telefone (Disparo)",
        "Tel_8dig": "Tel. (identificador)",
        "Tel_Limpo_Vendas": "Tel. Limpo (Vendas)",
        "Tel_8dig_Vendas": "Tel. 8 dígitos (Vendas)",
        "Tel_Limpo_Kommo": "Tel. Limpo (Kommo)",
        "Tel_8dig_Kommo": "Tel. 8 dígitos (Kommo)",
        "Data_Disparo": "Data do Disparo",
        "Data_Venda": "Data da Venda",
        "Dias_Após_Disparo": "Dias após o Disparo",
        "Situacao_Venda": "Situação do Registro",
        "Origem": "Origem da Venda",
        "Criterio_Match": "Como encontrou",
    })


def _write_guide_sheet(ws, summary_data: dict):
    """Escreve a aba 'Como Ler' com explicações claras para o usuário final."""
    PURPLE = "7B2FBE"
    DARK   = "1C1C1E"
    WHITE  = "FFFFFF"
    GRAY   = "F5F5F7"
    GREEN  = "27AE60"
    ORANGE = "E67E22"
    RED    = "C0392B"
    BLUE   = "2471A3"

    def _title(row, text, hex_bg=PURPLE):
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, text)
        c.font = Font(color=WHITE, bold=True, size=13)
        c.fill = PatternFill("solid", fgColor=hex_bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 28

    def _row(row, label, desc, label_bold=False):
        ws.merge_cells(f"B{row}:D{row}")
        a = ws.cell(row, 1, label)
        a.font = Font(bold=label_bold, size=10)
        a.fill = PatternFill("solid", fgColor=GRAY)
        a.alignment = Alignment(vertical="center", indent=1)
        b = ws.cell(row, 2, desc)
        b.font = Font(size=10)
        b.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 20

    def _color_legend(row, color_hex, label, meaning):
        ws.cell(row, 1, "  ██").font = Font(color=color_hex, size=14, bold=True)
        ws.cell(row, 1).fill = PatternFill("solid", fgColor=GRAY)
        ws.merge_cells(f"B{row}:C{row}")
        ws.cell(row, 2, label).font = Font(bold=True, size=10)
        ws.cell(row, 4, meaning).font = Font(size=10)
        ws.row_dimensions[row].height = 18

    r = 1
    # ── Cabeçalho ─────────────────────────────────────────────────────
    ws.merge_cells(f"A{r}:D{r}")
    hdr = ws.cell(r, 1, "📋  PROC AURE — GUIA DE LEITURA")
    hdr.font = Font(color=WHITE, bold=True, size=16)
    hdr.fill = PatternFill("solid", fgColor=PURPLE)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r].height = 40
    r += 1

    ws.merge_cells(f"A{r}:D{r}")
    sub = ws.cell(r, 1, "Desenvolvido por João  ·  Aure Digital  ·  Proc Aure")
    sub.font = Font(italic=True, size=10, color="888888")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 18
    r += 2

    # ── O que é este relatório ────────────────────────────────────────
    _title(r, "O QUE É ESTE RELATÓRIO?", DARK); r += 1
    ws.merge_cells(f"A{r}:D{r}")
    ws.cell(r, 1, (
        "Este relatório cruza sua planilha de vendas com o Kommo CRM para identificar "
        "quantas vendas vieram do tráfego pago e/ou de campanhas de disparo (WhatsApp)."
    )).font = Font(size=10)
    ws.cell(r, 1).alignment = Alignment(wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 30
    r += 2

    # ── Abas do relatório ─────────────────────────────────────────────
    _title(r, "ABAS DESTE ARQUIVO — O QUE CADA UMA CONTÉM", DARK); r += 1
    sheets_info = [
        ("📋 Como Ler",              "Esta aba. Leia primeiro."),
        ("📊 Resumo",                "Números principais: conversões, receita, taxas. Comece aqui."),
        ("✅ Vendas — Tráfego",      "Leads que vieram do tráfego pago E fizeram uma compra."),
        ("📣 Vendas — Disparo",      "Leads que receberam disparo E fizeram uma compra dentro de 30 dias."),
        ("🔍 Duplicatas",            "Registros duplicados na planilha de vendas (se houver)."),
        ("📋 Todos os Leads",        "Todos os leads do Kommo, com coluna 'Comprou?' para filtrar."),
        ("📦 Vendas (dados brutos)", "Sua planilha de vendas original, processada."),
        ("🗂️ Kommo (dados brutos)",  "Sua planilha do Kommo, processada."),
    ]
    for name, desc in sheets_info:
        _row(r, name, desc, label_bold=True); r += 1
    r += 1

    # ── Colunas importantes ───────────────────────────────────────────
    _title(r, "COLUNAS IMPORTANTES — O QUE CADA UMA SIGNIFICA", DARK); r += 1
    cols_info = [
        ("Comprou?",             "SIM = lead fez uma compra | NÃO = não encontramos venda para este lead"),
        ("É Lead de Tráfego?",   "SIM = lead tem a tag de tráfego pago no Kommo"),
        ("Origem da Venda",      "Tráfego / Disparo / Tráfego + Disparo (foi impactado pelos dois)"),
        ("Tel. (8 dígitos)",     "Últimos 8 dígitos do telefone — usados para comparar as listas"),
        ("Dias após o Disparo",  "Quantos dias após o disparo a venda aconteceu (máx. 30 dias conta)"),
        ("Situação do Registro", "Única / Multi-compra / DUPLICATA — análise de repetição nas vendas"),
    ]
    for col, desc in cols_info:
        _row(r, col, desc); r += 1
    r += 1

    # ── Legenda de cores ──────────────────────────────────────────────
    _title(r, "LEGENDA DE CORES NAS TABELAS", DARK); r += 1
    colors = [
        (GREEN,  "Verde",   "Venda confirmada (Comprou? = SIM)"),
        (RED,    "Vermelho","Não comprou (Comprou? = NÃO) | Duplicata"),
        (BLUE,   "Azul",    "Lead de tráfego pago | Multi-compra"),
        (ORANGE, "Laranja", "Atenção: mesmo dia / sem data confirmada"),
    ]
    for hex_c, name, meaning in colors:
        _color_legend(r, hex_c, name, meaning); r += 1
    r += 1

    # ── Como interpretar ──────────────────────────────────────────────
    _title(r, "COMO INTERPRETAR OS RESULTADOS", DARK); r += 1
    tips = [
        ("Tráfego puro",      "Lead veio de anúncio e comprou — conversão de tráfego."),
        ("Disparo puro",      "Lead recebeu WhatsApp e comprou em até 30 dias — conversão de disparo."),
        ("Tráfego + Disparo", "Lead veio de anúncio MAS só comprou depois do disparo. A equipe decide a quem atribuir."),
        ("Duplicata",         "Mesmo cliente, mesma data, mesmo valor — provável lançamento duplo no sistema. Verificar."),
        ("Multi-compra",      "Mesmo número, datas diferentes — cliente recorrente. Contar normalmente."),
    ]
    for label, desc in tips:
        _row(r, label, desc, label_bold=True); r += 1
    r += 1

    # ── Resumo de números ─────────────────────────────────────────────
    if summary_data:
        _title(r, "RESUMO RÁPIDO DOS NÚMEROS", "FF6B35"); r += 1
        for label, val in summary_data.items():
            _row(r, label, str(val)); r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40


def _sum_result_vcol(df_res: Optional[pd.DataFrame], value_col: Optional[str]) -> float:
    """Soma [Venda] {value_col} no DataFrame de resultado. Retorna 0.0 se não disponível."""
    if not value_col or df_res is None or len(df_res) == 0:
        return 0.0
    vcol = f"[Venda] {value_col}"
    if vcol not in df_res.columns:
        return 0.0
    return float(df_res[vcol].apply(parse_value).sum())


def _brl(value) -> str:
    """Formata float como R$ brasileiro."""
    try:
        return f"R$ {float(value):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (TypeError, ValueError):
        return "—"


def _add_total_row(ws, df: pd.DataFrame, value_col_name: str = ""):
    """Adiciona linha de TOTAL somando APENAS a coluna de valor identificada."""
    if len(df) == 0 or not value_col_name or value_col_name not in df.columns:
        return
    last_data_row = len(df) + 2
    total_row = last_data_row + 1
    BOLD_FILL = PatternFill("solid", fgColor="1C1C1E")

    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(total_row, c)
        cell.fill = BOLD_FILL
        if c == 1:
            cell.value = "TOTAL"
            cell.font = Font(color="FFFFFF", bold=True, size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == value_col_name:
            try:
                total = df[col].apply(parse_value).sum()
                cell.value = _brl(total)
                cell.font = Font(bold=True, size=12, color="FF6B35")
                cell.alignment = Alignment(horizontal="center")
            except Exception:
                pass
    ws.row_dimensions[total_row].height = 24


def build_excel(
    ds: pd.DataFrame,
    dk: pd.DataFrame,
    df_result: pd.DataFrame,
    df_full: pd.DataFrame,
    df_disparo_result: Optional[pd.DataFrame] = None,
    df_dup_analysis: Optional[pd.DataFrame] = None,
    sales_value_col: Optional[str] = None,
    df_breakdown: Optional[pd.DataFrame] = None,
) -> bytes:
    wb = Workbook()

    total_traffic = int((df_full["É_Tráfego"] == "SIM").sum()) if "É_Tráfego" in df_full.columns else 0
    total_sales   = len(ds)
    confirmed     = len(df_result)
    conv_rate     = f"{confirmed/total_traffic*100:.1f}%" if total_traffic > 0 else "—"

    # Prepara resumo para a aba guia
    guide_summary = {
        "Vendas carregadas": total_sales,
        "Leads no Kommo": len(dk),
        "Leads de tráfego": total_traffic,
        "Conversões de tráfego": confirmed,
        "Taxa de conversão (tráfego)": conv_rate,
    }
    if df_disparo_result is not None and len(df_disparo_result) > 0:
        d_conv = int((df_disparo_result["Venda_Confirmada"] == "SIM").sum())
        d_tot  = len(df_disparo_result)
        guide_summary["Conversões de disparo"] = d_conv
        guide_summary["Taxa de conversão (disparo)"] = f"{d_conv/d_tot*100:.1f}%" if d_tot > 0 else "—"

    # ── Aba 1: Como Ler (primeira aba — lida antes de tudo) ───────────
    ws_guide = wb.active
    ws_guide.title = "📋 Como Ler"
    _write_guide_sheet(ws_guide, guide_summary)

    # ── Aba 2: Resumo ─────────────────────────────────────────────────
    ws_res = wb.create_sheet("📊 Resumo")
    ws_res.merge_cells("A1:C1")
    t = ws_res.cell(1, 1, "RESUMO DO PROC AURE")
    t.font = Font(color="FFFFFF", bold=True, size=14)
    t.fill = PatternFill("solid", fgColor="7B2FBE")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_res.row_dimensions[1].height = 36

    summary_rows = [
        ("Vendas carregadas", total_sales),
        ("Leads no Kommo", len(dk)),
        ("— TRÁFEGO —", ""),
        ("Leads com tag de tráfego", total_traffic),
        ("Conversões confirmadas", confirmed),
        ("Taxa de conversão", conv_rate),
        ("→ Veja os valores na aba ✅ Vendas — Tráfego", ""),
    ]
    if df_disparo_result is not None and len(df_disparo_result) > 0:
        d_tot  = len(df_disparo_result)
        d_conv = int((df_disparo_result["Venda_Confirmada"] == "SIM").sum())
        d_rate = f"{d_conv/d_tot*100:.1f}%" if d_tot > 0 else "—"
        summary_rows += [
            ("— DISPARO —", ""),
            ("Leads de disparo analisados", d_tot),
            ("Conversões confirmadas", d_conv),
            ("Taxa de conversão", d_rate),
            ("→ Veja os valores na aba 📣 Vendas — Disparo", ""),
        ]
    if df_dup_analysis is not None and "Situacao_Venda" in df_dup_analysis.columns:
        n_dup   = int((df_dup_analysis["Situacao_Venda"].str.startswith("DUPLICATA")).sum())
        n_multi = int((df_dup_analysis["Situacao_Venda"].str.startswith("Multi")).sum())
        summary_rows += [
            ("— QUALIDADE DOS DADOS —", ""),
            ("Duplicatas suspeitas", n_dup),
            ("Multi-compras (cliente recorrente)", n_multi),
        ]
    for i, (label, val) in enumerate(summary_rows, 2):
        lc = ws_res.cell(i, 1, label)
        lc.font = Font(bold=True, size=11)
        if str(label).startswith("—"):
            lc.fill = PatternFill("solid", fgColor="7B2FBE")
            lc.font = Font(bold=True, size=11, color="FFFFFF")
        vc = ws_res.cell(i, 2, val)
        vc.font = Font(bold=True, size=13, color="FF6B35")
        vc.alignment = Alignment(horizontal="center")
        ws_res.row_dimensions[i].height = 26
    ws_res.column_dimensions["A"].width = 46
    ws_res.column_dimensions["B"].width = 22

    # Coluna de valor renomeada (para _add_total_row)
    _vcol = f"[Venda] {sales_value_col}" if sales_value_col else ""

    # ── Aba: Mês a Mês (por aba) ──────────────────────────────────────
    if df_breakdown is not None and len(df_breakdown) > 0:
        ws_bd = wb.create_sheet("📅 Por Lista-Mês")
        _write_sheet(ws_bd, df_breakdown, "VENDAS POR LISTA / MÊS (Kommo × cada lista ou aba)", "7B2FBE")

    # ── Aba 3: Vendas — Tráfego ───────────────────────────────────────
    ws3 = wb.create_sheet("✅ Vendas — Tráfego")
    if len(df_result) == 0:
        ws3.cell(1, 1, "Nenhuma venda de tráfego confirmada.").font = Font(italic=True, color="888888")
    else:
        df_result_renamed = _rename_result_cols(df_result)
        _write_sheet(ws3, df_result_renamed,
                     f"VENDAS DO TRÁFEGO PAGO — {confirmed} CONVERSÕES CONFIRMADAS", "27AE60")
        _add_total_row(ws3, df_result_renamed, _vcol)

    # ── Aba 4: Vendas — Disparo ───────────────────────────────────────
    all_sheets = [ws_guide, ws_res, ws3]
    if df_disparo_result is not None and len(df_disparo_result) > 0:
        ws_disp = wb.create_sheet("📣 Vendas — Disparo")
        d_conv  = int((df_disparo_result["Venda_Confirmada"] == "SIM").sum())
        df_disp_renamed = _rename_result_cols(df_disparo_result)
        _write_sheet(
            ws_disp, df_disp_renamed,
            f"VENDAS DO DISPARO — {d_conv} CONVERSÕES CONFIRMADAS",
            "8E44AD",
            ["Data do Disparo", "Data da Venda", "Dias após o Disparo"],
        )
        _add_total_row(ws_disp, df_disp_renamed, _vcol)
        all_sheets.append(ws_disp)

    # ── Aba 5: Duplicatas ─────────────────────────────────────────────
    if df_dup_analysis is not None and "Situacao_Venda" in df_dup_analysis.columns:
        df_non_unique = df_dup_analysis[df_dup_analysis["Situacao_Venda"] != "Única"]
        if len(df_non_unique) > 0:
            n_dup   = int((df_dup_analysis["Situacao_Venda"].str.startswith("DUPLICATA")).sum())
            n_multi = int((df_dup_analysis["Situacao_Venda"].str.startswith("Multi")).sum())
            ws_dup  = wb.create_sheet("🔍 Duplicatas")
            _write_sheet(
                ws_dup, _rename_result_cols(df_non_unique),
                f"DUPLICATAS ({n_dup}) E MULTI-COMPRAS ({n_multi}) — VERIFIQUE",
                "922B21", ["Situação do Registro"],
            )
            all_sheets.append(ws_dup)

    # ── Aba 6: Todos os Leads ─────────────────────────────────────────
    ws4 = wb.create_sheet("📋 Todos os Leads")
    _write_sheet(ws4, _rename_result_cols(df_full),
                 "TODOS OS LEADS DO KOMMO — FILTRE PELA COLUNA 'COMPROU?'", "6C3483")
    all_sheets.append(ws4)

    # ── Aba 7: Dados Brutos — Vendas ──────────────────────────────────
    ws1 = wb.create_sheet("📦 Vendas (bruto)")
    _write_sheet(ws1, _rename_result_cols(ds),
                 "PLANILHA DE VENDAS PROCESSADA", "FF6B35",
                 ["Tel. Limpo (Vendas)", "Tel. 8 dígitos (Vendas)"])
    all_sheets.append(ws1)

    # ── Aba 8: Dados Brutos — Kommo ───────────────────────────────────
    ws2 = wb.create_sheet("🗂️ Kommo (bruto)")
    _write_sheet(ws2, _rename_result_cols(dk),
                 "PLANILHA KOMMO PROCESSADA", "2980B9",
                 ["Tel. Limpo (Kommo)", "Tel. 8 dígitos (Kommo)"])
    all_sheets.append(ws2)

    # ── Rodapé em todas as abas ───────────────────────────────────────
    for ws in all_sheets:
        r = ws.max_row + 2
        ws.cell(r, 1, "Desenvolvido por João  ·  Proc Aure  ·  Aure Digital").font = Font(
            italic=True, color="AAAAAA", size=9
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Inteligência Claude (opcional) ─────────────────────────────────────────────
# A ferramenta funciona 100% sem chave de API — a IA é uma camada extra que
# confere colunas, classifica tags, revisa matches duvidosos e explica resultados.

_CLAUDE_MODEL = "claude-opus-4-8"


def _get_api_key() -> Optional[str]:
    """Chave da API da Anthropic: st.secrets → variável de ambiente → sidebar."""
    key = None
    try:
        key = st.secrets.get("ANTHROPIC_API_KEY")
    except Exception:
        pass
    return key or os.environ.get("ANTHROPIC_API_KEY") or st.session_state.get("_ai_key")


def _mask_pii(v) -> str:
    """Mascara o miolo de sequências com 6+ dígitos (telefone/CPF) antes de
    enviar amostras para a API — o suficiente para reconhecer o TIPO da coluna."""
    s = str(v)
    out, runs = list(s), []
    start = None
    for i, ch in enumerate(s + " "):
        if ch.isdigit():
            if start is None:
                start = i
        elif start is not None:
            runs.append((start, i))
            start = None
    for a, b in runs:
        if b - a >= 6:
            for i in range(a + 2, b - 2):
                out[i] = "•"
    return "".join(out)


class _AIError(Exception):
    """Erro amigável de IA. Levantado dentro das funções cacheadas para que o
    st.cache_data NÃO guarde falhas (ex.: chave inválida corrigida depois)."""


def _ai_safe(fn, *args) -> tuple:
    """Chama uma função de IA cacheada devolvendo (resultado, erro_amigável)."""
    try:
        return fn(*args), None
    except _AIError as e:
        return None, str(e)


def _ai_call(system: str, user_text: str, schema: Optional[dict] = None,
             max_tokens: int = 4000, api_key: Optional[str] = None):
    """Uma chamada ao Claude. Retorna o resultado ou levanta _AIError.
    Com schema → dict validado; sem schema → texto."""
    key = api_key or _get_api_key()
    if not key:
        raise _AIError("Sem chave de API configurada.")
    try:
        import anthropic
    except ImportError:
        raise _AIError("O pacote de IA ainda está sendo instalado no servidor — "
                       "aguarde 1-2 minutos e recarregue a página.")
    try:
        client = anthropic.Anthropic(api_key=key)
        kwargs: dict = {}
        if schema is not None:
            kwargs["output_config"] = {"format": {"type": "json_schema", "schema": schema}}
        resp = client.messages.create(
            model=_CLAUDE_MODEL,
            max_tokens=max_tokens,
            thinking={"type": "adaptive"},
            system=system,
            messages=[{"role": "user", "content": user_text}],
            **kwargs,
        )
        if resp.stop_reason == "refusal":
            raise _AIError("O Claude recusou esta solicitação.")
        text = next((b.text for b in resp.content if b.type == "text"), "")
        if schema is not None:
            return json.loads(text)
        return text
    except _AIError:
        raise
    except anthropic.AuthenticationError:
        raise _AIError("Chave de API inválida — confira em console.anthropic.com.")
    except anthropic.RateLimitError:
        raise _AIError("Limite de requisições atingido — aguarde um minuto e tente de novo.")
    except anthropic.APIConnectionError:
        raise _AIError("Sem conexão com a API da Anthropic — verifique a internet.")
    except anthropic.APIStatusError as e:
        raise _AIError(f"Erro da API ({e.status_code}) — tente novamente.")
    except Exception as e:
        raise _AIError(f"Erro inesperado na IA: {type(e).__name__}.")


def _ai_payload_planilha(df: pd.DataFrame, rotulo: str, n_amostras: int = 5) -> str:
    """Nome das colunas + amostras mascaradas — o que o Claude precisa para mapear."""
    amostras = {}
    for c in list(df.columns)[:80]:
        vals = df[c].dropna().astype(str).head(n_amostras).tolist()
        amostras[str(c)] = [_mask_pii(v)[:40] for v in vals]
    return json.dumps(
        {"planilha": rotulo, "total_linhas": len(df), "colunas_e_amostras": amostras},
        ensure_ascii=False,
    )


_SCHEMA_COLUNAS = {
    "type": "object",
    "properties": {
        "telefone": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "nome": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "valor": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "data": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "tags": {"anyOf": [{"type": "string"}, {"type": "null"}]},
        "observacao": {"type": "string"},
    },
    "required": ["telefone", "nome", "valor", "data", "tags", "observacao"],
    "additionalProperties": False,
}


@st.cache_data(show_spinner=False)
def ai_mapear_colunas(payload: str):
    """Claude aponta qual coluna é telefone/nome/valor/data/tags."""
    system = (
        "Você analisa planilhas brasileiras de vendas e exports do CRM Kommo. "
        "Dado o nome das colunas e amostras (telefones mascarados com •), identifique a MELHOR coluna para cada papel: "
        "telefone do cliente (a mais preenchida com números de telefone), nome do cliente, "
        "valor da venda em R$ (não quantidade, não código), data da venda (a que varia por linha; "
        "ignore datas de período do relatório, que são constantes, e colunas quase vazias tipo "
        "data de emissão no rodapé — nesses casos responda null) e tags/etiquetas. "
        "Use null quando não existir coluna para o papel. Responda com o nome EXATO da coluna. "
        "Em 'observacao', avise em 1-2 frases sobre problemas que notou (ex.: telefones sem DDD, "
        "coluna de valor em formato americano, datas constantes)."
    )
    return _ai_call(system, payload, _SCHEMA_COLUNAS)


_SCHEMA_TAGS = {
    "type": "object",
    "properties": {
        "trafego_incluir": {"type": "string"},
        "trafego_excluir": {"type": "string"},
        "disparo_incluir": {"type": "string"},
        "disparo_excluir": {"type": "string"},
        "explicacao": {"type": "string"},
    },
    "required": ["trafego_incluir", "trafego_excluir", "disparo_incluir",
                 "disparo_excluir", "explicacao"],
    "additionalProperties": False,
}


@st.cache_data(show_spinner=False)
def ai_sugerir_tags(payload: str):
    """Claude classifica as tags do Kommo em tráfego pago × disparo × outras."""
    system = (
        "Você conhece o CRM Kommo usado por agências de tráfego pago no Brasil. "
        "Receberá a lista de tags de leads com a contagem de cada uma. Classifique: "
        "quais identificam leads de TRÁFEGO PAGO (ex.: TRAFEGO, Tráfego, ads) e quais identificam "
        "DISPAROS de WhatsApp (ex.: 'DISPARO NAMORADOS 09/06/26', 'DISPARO LIVE'). "
        "Tags de origem orgânica (INSTAGRAM, indicação), de vendedoras (nomes próprios), de importação "
        "(IMPORT LEADS) e de teste NÃO são tráfego nem disparo. "
        "Responda com palavras-chave em minúsculas, separadas por vírgula, prontas para busca por "
        "substring sem acento (ex.: 'trafego' casa 'TRÁFEGO' e 'Tráfego'). Prefira UMA palavra-chave "
        "curta que cubra o grupo todo; use o campo excluir para separar falsos positivos "
        "(ex.: excluir 'organico' quando existir 'Tráfego Orgânico'). Deixe campos vazios se não houver. "
        "Em 'explicacao', 1-3 frases sobre o que você encontrou."
    )
    return _ai_call(system, payload, _SCHEMA_TAGS)


_SCHEMA_REVISAO = {
    "type": "object",
    "properties": {
        "veredictos": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "indice": {"type": "integer"},
                    "veredito": {"type": "string",
                                 "enum": ["mesma_pessoa", "incerto", "pessoas_diferentes"]},
                    "motivo": {"type": "string"},
                },
                "required": ["indice", "veredito", "motivo"],
                "additionalProperties": False,
            },
        },
    },
    "required": ["veredictos"],
    "additionalProperties": False,
}


@st.cache_data(show_spinner=False)
def ai_revisar_matches(payload: str):
    """Claude dá veredicto nos matches ⚠️ feitos por nome (sem telefone batendo)."""
    system = (
        "Você revisa cruzamentos de clientes brasileiros feitos por NOME (o telefone não bateu). "
        "Para cada par (nome no CRM × nome na lista de vendas), diga se é a mesma pessoa. "
        "Considere: abreviações (Mª, Jr), sobrenomes de casada, apelidos comuns (Zé/José, Bia/Beatriz), "
        "ordem dos sobrenomes. Nomes muito comuns com pouca informação (ex.: 'Maria Silva' × 'Maria Silva') "
        "são 'incerto' — homônimos são frequentes. Só responda 'mesma_pessoa' quando o nome for "
        "distintivo o suficiente. 'motivo' em até 12 palavras."
    )
    return _ai_call(system, payload, _SCHEMA_REVISAO)


_CHAT_SYSTEM = (
    "Você é o SUPERVISOR da Proc Aure, conversando DENTRO da ferramenta com o usuário sobre um "
    "processamento que acabou de rodar. O primeiro bloco da conversa traz o contexto em JSON: "
    "resumo dos números, tratamento aplicado em cada planilha, colunas e mapeamento usado, como "
    "o cruzamento funciona, conversões e amostras de quem NÃO converteu. "
    "Além disso você tem a ferramenta consultar_planilha, que acessa as TABELAS REAIS deste "
    "processamento (vendas, kommo, conversões, cruzamento completo) — use-a sempre que a "
    "resposta depender de um dado que não está no contexto: conferir uma venda específica, "
    "um telefone, um lead, quantas linhas têm certa tag. Nunca diga que não tem acesso à "
    "planilha: consulte. "
    "DOUTRINA DO CRUZAMENTO — telefone é a identidade: o PROCV bate telefone com telefone e "
    "depois olha a tag; TODAS as colunas de telefone dos dois lados são varridas, então um "
    "cliente com 3 números cadastrados casa por qualquer um deles (e conta UMA vez). Match "
    "por nome é exceção de último recurso, sempre ⚠️. Ao investigar um cliente, consulte a "
    "linha dele na tabela 'vendas', anote TODOS os telefones (todas as colunas) e busque "
    "CADA número no kommo antes de concluir que não casou — nunca pare no primeiro número "
    "nem se apoie em busca por nome quando houver telefone para conferir. "
    "Critérios de match: Telefone; 'Col. alt.' = telefone achado em outra "
    "coluna; 'Nome ⚠️' = casou só pelo nome, merece conferência; janela = venda até 30 dias "
    "após a data do disparo; bloqueado por DDD = mesmo final de número em cidades diferentes. "
    "Aja como supervisor: se notar problema de tratamento (coluna errada, telefones inválidos, "
    "tag mal escolhida, datas fora do período), aponte e explique como corrigir e reprocessar. "
    "Quando o usuário corrigir algo (ex.: 'esses dois são a mesma pessoa'), aceite, use no "
    "resto da conversa e registre numa linha final começando com '📌 Anotado:'. "
    "Não invente números — confira na tabela antes de afirmar. Responda em português, direto "
    "e sem jargão."
)

_CHAT_TOOL = {
    "name": "consultar_planilha",
    "description": (
        "Consulta as tabelas reais DESTE processamento, linha a linha. Tabelas: "
        "'vendas' (planilha de vendas tratada), 'kommo' (leads do CRM), "
        "'conversoes_trafego' (vendas atribuídas ao tráfego), "
        "'todos_os_leads_cruzados' (cada lead do Kommo com É_Tráfego e Venda_Confirmada), "
        "'disparo' (leads de disparo com Venda_Confirmada). "
        "Use 'buscar' para filtrar por nome, telefone (qualquer formato) ou tag."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "tabela": {"type": "string",
                       "enum": ["vendas", "kommo", "conversoes_trafego",
                                "todos_os_leads_cruzados", "disparo"]},
            "buscar": {"type": "string",
                       "description": "Texto ou telefone a procurar em qualquer coluna. "
                                      "Vazio = primeiras linhas da tabela."},
            "colunas": {"type": "array", "items": {"type": "string"},
                        "description": "Colunas a retornar (vazio = principais)."},
            "max_linhas": {"type": "integer", "description": "Máximo de linhas (até 40)."},
        },
        "required": ["tabela"],
    },
}


def _exec_consulta(inp: dict) -> str:
    """Executa a consulta do chat nas tabelas guardadas na sessão."""
    try:
        dfs = st.session_state.get("_ai_dfs") or {}
        tab = str(inp.get("tabela", ""))
        df = dfs.get(tab)
        if df is None or len(df) == 0:
            return json.dumps({
                "erro": f"tabela '{tab}' vazia ou inexistente neste processamento",
                "tabelas_disponiveis": [k for k, v in dfs.items()
                                        if v is not None and len(v) > 0],
            }, ensure_ascii=False)
        out = df
        buscar = str(inp.get("buscar") or "").strip()
        if buscar:
            digits = re.sub(r"\D", "", buscar)
            if len(digits) >= 6:
                sub = digits[-8:]
                mask = out.apply(lambda r: any(
                    sub in re.sub(r"\D", "", str(v)) for v in r), axis=1)
            else:
                b = _normalize(buscar)
                mask = out.apply(lambda r: any(
                    b in _normalize(str(v)) for v in r), axis=1)
            out = out[mask]
        cols = [c for c in (inp.get("colunas") or []) if c in out.columns]
        if not cols:
            cols = [c for c in out.columns if not str(c).startswith("_")][:14]
        n = min(int(inp.get("max_linhas") or 20), 40)
        recs = (out[cols].head(n).astype(str)
                .apply(lambda s: s.str.slice(0, 80)).to_dict("records"))
        return json.dumps({
            "total_de_linhas_encontradas": int(len(out)),
            "mostrando": len(recs),
            "colunas_disponiveis": [str(c) for c in df.columns][:60],
            "linhas": recs,
        }, ensure_ascii=False)
    except Exception as e:
        return json.dumps({"erro": f"{type(e).__name__}: {e}"}, ensure_ascii=False)


def _ai_chat_reply(ctx: str, historico: list) -> tuple:
    """Um turno do chat sobre o resultado. Retorna (texto, erro_amigável).
    O contexto entra como primeiro bloco com cache — os turnos seguintes
    reaproveitam o prefixo e custam centavos."""
    key = _get_api_key()
    if not key:
        return None, "Sem chave de API configurada."
    try:
        import anthropic
    except ImportError:
        return None, ("O pacote de IA ainda está sendo instalado no servidor — "
                      "aguarde 1-2 minutos e recarregue a página.")
    try:
        client = anthropic.Anthropic(api_key=key)
        msgs = [
            {"role": "user", "content": [{
                "type": "text",
                "text": "CONTEXTO DO PROCESSAMENTO (JSON gerado pela Proc Aure):\n" + ctx,
                "cache_control": {"type": "ephemeral"},
            }]},
            {"role": "assistant",
             "content": "Recebi o contexto do processamento. Pode perguntar sobre o resultado."},
        ]
        for m in historico:
            msgs.append({"role": m["role"], "content": m["content"]})
        # Loop agêntico: o Claude consulta as tabelas reais quantas vezes precisar
        for _ in range(6):
            resp = client.messages.create(
                model=_CLAUDE_MODEL,
                max_tokens=3000,
                thinking={"type": "adaptive"},
                system=_CHAT_SYSTEM,
                tools=[_CHAT_TOOL],
                messages=msgs,
            )
            if resp.stop_reason == "refusal":
                return None, "O Claude recusou esta solicitação."
            if resp.stop_reason != "tool_use":
                return next((b.text for b in resp.content if b.type == "text"), ""), None
            msgs.append({"role": "assistant", "content": resp.content})
            results = []
            for b in resp.content:
                if b.type == "tool_use":
                    results.append({
                        "type": "tool_result",
                        "tool_use_id": b.id,
                        "content": _exec_consulta(dict(b.input)),
                    })
            msgs.append({"role": "user", "content": results})
        return None, "A consulta ficou longa demais — tente uma pergunta mais específica."
    except anthropic.AuthenticationError:
        return None, "Chave de API inválida — confira em console.anthropic.com."
    except anthropic.RateLimitError:
        return None, "Limite de requisições atingido — aguarde um minuto e tente de novo."
    except anthropic.APIConnectionError:
        return None, "Sem conexão com a API da Anthropic — verifique a internet."
    except anthropic.APIStatusError as e:
        return None, f"Erro da API ({e.status_code}) — tente novamente."
    except Exception as e:
        return None, f"Erro inesperado na IA: {type(e).__name__}."


@st.cache_data(show_spinner=False)
def ai_explicar_resultados(payload: str):
    """Resumo executivo: o que os números significam e por que (ex.: 0 conversões)."""
    system = (
        "Você é o analista da Proc Aure, ferramenta que cruza listas de vendas com exports do Kommo "
        "para atribuir vendas a tráfego pago e a disparos de WhatsApp. Receberá um JSON com os números "
        "de um processamento. Escreva um resumo executivo em português, markdown, máximo 10 linhas: "
        "1) o resultado principal em linguagem de dono de negócio; 2) SE houver algo estranho "
        "(0 conversões, datas de disparo fora do período das vendas, muitas vendas sem telefone, "
        "tag que não casou com nada), explique O PORQUÊ com os dados e diga o que fazer; "
        "3) uma recomendação prática. Não invente números que não estão no JSON. "
        "Não use jargão técnico. Seja direto — sem introduções."
    )
    return _ai_call(system, payload, max_tokens=2000)


# ── Interface ──────────────────────────────────────────────────────────────────

def _phone_preview(df: pd.DataFrame, col: Optional[str], is_auto: bool = False) -> str:
    """Prévia da coluna de telefone ANTES de rodar: quantos valores parecem telefone.
    Avisa quando a coluna provavelmente está errada (preview/guardrail)."""
    if col is None or col not in df.columns:
        return ""
    sample = df[col].dropna().astype(str).head(50)
    n = len(sample)
    if n == 0:
        return "⚠️ coluna sem dados"
    hits = int(sample.apply(_looks_like_phone).sum())
    prefix = "✨ Auto-detectado · " if is_auto else ""
    if hits / n >= 0.6:
        return f"{prefix}✅ {hits}/{n} amostrados parecem telefone"
    if hits / n >= 0.2:
        return f"{prefix}⚠️ só {hits}/{n} parecem telefone — confira a coluna"
    return f"{prefix}⚠️ {hits}/{n} parecem telefone — provavelmente NÃO é a coluna certa"


def _show_treatment_notes(df: Optional[pd.DataFrame], origem: str) -> None:
    """Exibe o bloco '🧠 O que a ferramenta tratou' a partir das notas em df.attrs."""
    if df is None:
        return
    notes = df.attrs.get("treatment_notes", [])
    if not notes:
        return
    with st.expander(f"🧠 O que a ferramenta tratou — {origem}", expanded=True):
        for n in notes:
            st.markdown(f"• {n}")


def _load_multi(files, combiner=combine_kommo_sources) -> Optional[pd.DataFrame]:
    """Carrega e COMBINA uma lista de arquivos (funis do Kommo ou listas de
    vendas). Cada arquivo é lido por inteiro (todas as abas) e combinado."""
    dfs, names = [], []
    for f in files:
        b = f.read()
        sheets = _get_sheets_cached(b, f.name)
        df, _info = _load_cached(b, f.name, tuple(sheets or ["__csv__"]))
        if df is not None and len(df) > 0:
            dfs.append(df)
            names.append(f.name)
    return combiner(dfs, names)


st.markdown("""
<div class="hero-wrap">
  <div class="hero-badge">Aure Digital</div>
  <h1 class="hero-title" translate="no">Proc&nbsp;Aure</h1>
  <p class="hero-subtitle">Análise inteligente de conversão · Tráfego &amp; Disparo</p>
</div>
""", unsafe_allow_html=True)
st.divider()

# ── Sidebar: Inteligência Claude (opcional) ────────────────────────────────────
with st.sidebar:
    st.markdown("### 🤖 Inteligência Claude")
    # Copia o valor digitado para um slot DURÁVEL da sessão: o Streamlit apaga
    # o estado de widget que sai da tela — era o bug da chave "sumindo da
    # caixinha" (e do chat mudo) assim que a caixinha virava "Conectado".
    if st.session_state.get("_ai_key_input"):
        st.session_state["_ai_key"] = st.session_state["_ai_key_input"]
    if _get_api_key():
        st.caption("✅ Conectado — recursos de IA ativos.")
        if st.session_state.get("_ai_key") and st.button("🔁 Trocar chave"):
            st.session_state.pop("_ai_key", None)
            st.rerun()
    else:
        st.text_input(
            "Chave da API da Anthropic", type="password", key="_ai_key_input",
            help="Opcional — a ferramenta funciona normalmente sem IA. "
                 "Crie a chave em console.anthropic.com → API Keys. "
                 "No Streamlit Cloud, configure ANTHROPIC_API_KEY em "
                 "Settings → Secrets UMA vez para nunca mais precisar colar.",
        )
        if not _get_api_key():
            st.caption("Sem chave: a ferramenta roda normal, só sem os recursos 🤖.")
    _has_ai = bool(_get_api_key())
    ai_auto = st.toggle(
        "Análise automática ao processar", value=True, disabled=not _has_ai,
        help="Após cada processamento, o Claude explica os resultados e revisa "
             "os matches ⚠️ feitos por nome. Consome créditos da sua API "
             "(centavos por processamento).",
    )
    if _has_ai:
        st.caption("O Claude confere as planilhas **automaticamente** ao carregar, "
                   "sugere as tags no passo 2, analisa o resultado e responde no chat "
                   "consultando as tabelas reais — como um supervisor do cruzamento.")

# Selo de status da IA — sempre visível na página principal
if _has_ai:
    st.markdown(
        '<p style="text-align:center;font-size:.92rem;color:#30D158;margin-top:-.4rem">'
        '🤖 <b>Inteligência Claude ativa</b> — conferência de colunas e tags no passo 2 '
        '· análise executiva e revisão de matches nos resultados</p>',
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        '<p style="text-align:center;font-size:.92rem;color:rgba(235,235,245,.5);margin-top:-.4rem">'
        '🤖 Inteligência Claude <b>desligada</b> — cole sua chave da Anthropic '
        'na barra lateral (à esquerda) para ativar os recursos de IA</p>',
        unsafe_allow_html=True,
    )

# ── Passo 1: Upload ────────────────────────────────────────────────────────────
st.markdown('<div class="step-wrap"><div class="step-num">1</div><div class="step-text">Carregue as planilhas</div></div>', unsafe_allow_html=True)

col_left, col_right = st.columns(2)

with col_left:
    st.markdown("#### 📊 Planilha(s) de Vendas")
    sales_files = st.file_uploader(
        "Arraste ou clique — pode soltar VÁRIAS listas (ex.: uma por vendedora)",
        type=["xlsx", "xls", "csv", "xlsm"],
        key="sales_upload",
        accept_multiple_files=True,
        help="Planilha do cliente com nome, telefone e valor de compra. "
             "Se o cliente manda uma lista por vendedora, solte todas aqui: "
             "a ferramenta cruza cada uma com o Kommo e entrega o resultado "
             "POR LISTA e o total combinado (comprador repetido conta 1× no total).",
    )
    df_sales_raw: Optional[pd.DataFrame] = None
    if sales_files and len(sales_files) > 1:
        # Várias listas (ex.: uma por vendedora) → combina marcando a origem
        df_sales_raw = _load_multi(sales_files, combine_sales_sources)
        if df_sales_raw is not None:
            st.success(f"✅ {len(sales_files)} listas combinadas · "
                       f"{len(df_sales_raw):,} vendas · {len(df_sales_raw.columns)} colunas")
            st.caption("📑 O resultado sai separado por lista (seção 'Por Lista / Mês') "
                       "e com o total combinado.")
            _show_treatment_notes(df_sales_raw, "Vendas")
            with st.expander("Prévia"):
                st.dataframe(df_sales_raw.head(6), use_container_width=True)
    elif sales_files:
        sales_file = sales_files[0]
        sales_bytes = sales_file.read()  # lê UMA vez aqui
        sales_sheets = _get_sheets_cached(sales_bytes, sales_file.name)
        selected_sales_sheets = sales_sheets

        if len(sales_sheets) > 1:
            st.caption(f"📂 {len(sales_sheets)} planilhas encontradas neste arquivo")
            selected_sales_sheets = st.multiselect(
                "Selecione as planilhas a usar:",
                options=sales_sheets,
                default=sales_sheets,
                key="sales_sheets_select",
            )
            if not selected_sales_sheets:
                st.warning("Selecione ao menos uma planilha.")

        if selected_sales_sheets or not sales_sheets:
            df_sales_raw, sales_info = _load_cached(
                sales_bytes, sales_file.name, tuple(selected_sales_sheets or ["__csv__"])
            )
            if df_sales_raw is not None:
                for sheet, hrow in sales_info.items():
                    if hrow > 0:
                        st.info(f"📋 '{sheet}': cabeçalho detectado na linha {hrow + 1} — título(s) anteriores ignorados.")
                # Aviso sobre abas primárias vs secundárias
                sec = df_sales_raw.attrs.get("secondary_sheets", [])
                pri = df_sales_raw.attrs.get("primary_sheets", [])
                if sec:
                    st.warning(
                        f"🧠 **Análise inteligente de abas:** "
                        f"A aba **'{', '.join(sec)}'** não tem estrutura completa de vendas "
                        f"(sem coluna de data ou muitos campos sem nome). "
                        f"Ela será usada apenas para cruzamento de telefones. "
                        f"A análise principal (datas, valores, duplicatas) usa: **'{', '.join(pri)}'**."
                    )
                st.success(f"✅ {len(df_sales_raw):,} linhas · {len(df_sales_raw.columns)} colunas")
                _show_treatment_notes(df_sales_raw, "Vendas")
                with st.expander("Prévia"):
                    st.dataframe(df_sales_raw.head(6), use_container_width=True)

with col_right:
    st.markdown("#### 🗂️ Planilha(s) do Kommo")
    kommo_files = st.file_uploader(
        "Arraste ou clique — pode jogar VÁRIOS arquivos (um por funil)",
        type=["xlsx", "xls", "csv", "xlsm"],
        key="kommo_upload",
        accept_multiple_files=True,
        help="Export(s) do Kommo CRM com leads, telefones e tags. "
             "Se você tem vários funis, baixe um arquivo de cada e solte todos aqui — "
             "a ferramenta junta tudo num só e cruza com as vendas.",
    )
    df_kommo_raw: Optional[pd.DataFrame] = None
    if kommo_files:
        df_kommo_raw = _load_multi(kommo_files)
        if df_kommo_raw is not None:
            if len(kommo_files) > 1:
                st.success(f"✅ {len(kommo_files)} funis combinados · {len(df_kommo_raw):,} leads · {len(df_kommo_raw.columns)} colunas")
            else:
                st.success(f"✅ {len(df_kommo_raw):,} linhas · {len(df_kommo_raw.columns)} colunas")
            _show_treatment_notes(df_kommo_raw, "Kommo")
            with st.expander("Prévia"):
                st.dataframe(df_kommo_raw.head(6), use_container_width=True)

st.markdown("#### 📣 Planilha(s) do Kommo — Disparo *(opcional, se for arquivo separado do tráfego)*")
kommo_disparo_files = st.file_uploader(
    "Se o Kommo de disparo é separado do de tráfego, solte aqui (pode ser mais de um)",
    type=["xlsx", "xls", "csv", "xlsm"],
    key="kommo_disparo_upload",
    accept_multiple_files=True,
    help="Opcional. Se não carregar, o(s) arquivo(s) Kommo principal(is) servem para tráfego e disparo.",
)
df_kommo_disparo_raw: Optional[pd.DataFrame] = None
if kommo_disparo_files:
    df_kommo_disparo_raw = _load_multi(kommo_disparo_files)
    if df_kommo_disparo_raw is not None:
        _nf = len(kommo_disparo_files)
        st.success(f"✅ Kommo Disparo: {('%d funis · ' % _nf) if _nf > 1 else ''}{len(df_kommo_disparo_raw):,} leads · {len(df_kommo_disparo_raw.columns)} colunas")
        _show_treatment_notes(df_kommo_disparo_raw, "Kommo Disparo")
        with st.expander("Prévia"):
            st.dataframe(df_kommo_disparo_raw.head(6), use_container_width=True)

st.divider()

# ── Passo 2: Configuração ──────────────────────────────────────────────────────
if df_sales_raw is not None and df_kommo_raw is not None:
    st.markdown('<div class="step-wrap"><div class="step-num">2</div><div class="step-text">Configure as colunas</div></div>', unsafe_allow_html=True)

    auto_sp = detect_phone_col(df_sales_raw)
    auto_kp = detect_phone_col(df_kommo_raw)
    auto_kt = detect_tag_col(df_kommo_raw)
    auto_vl = detect_value_col(df_sales_raw)
    auto_sn = detect_name_col(df_sales_raw)
    auto_kn = detect_name_col(df_kommo_raw)

    # ── 🤖 IA participa do tratamento: confere as planilhas AUTOMATICAMENTE ──
    # Roda uma vez por par de arquivos (sucesso fica em cache); erro não gruda
    # e ganha botão de repetir.
    if _has_ai:
        _pv = _ai_payload_planilha(df_sales_raw, "lista de vendas")
        _pk = _ai_payload_planilha(df_kommo_raw, "export do Kommo")
        if st.session_state.get("_ai_cols_payloads") != (_pv, _pk):
            with st.spinner("🤖 Claude conferindo as planilhas..."):
                st.session_state["_ai_cols_v"] = _ai_safe(ai_mapear_colunas, _pv)
                st.session_state["_ai_cols_k"] = _ai_safe(ai_mapear_colunas, _pk)
            st.session_state["_ai_cols_payloads"] = (_pv, _pk)
        if (st.session_state.get("_ai_cols_v", (None, None))[1]
                or st.session_state.get("_ai_cols_k", (None, None))[1]):
            if st.button("🔁 Tentar a conferência das planilhas de novo"):
                st.session_state.pop("_ai_cols_payloads", None)
                st.rerun()
    _ai_v_res, _ai_v_err = st.session_state.get("_ai_cols_v", (None, None))
    _ai_k_res, _ai_k_err = st.session_state.get("_ai_cols_k", (None, None))
    for _ai_e in (_ai_v_err, _ai_k_err):
        if _ai_e:
            st.warning(f"🤖 {_ai_e}")

    def _ai_pick(res: Optional[dict], papel: str, df: pd.DataFrame, fallback):
        c = (res or {}).get(papel)
        return c if (c and c in df.columns) else fallback

    def _ai_conferido(res: Optional[dict], papel: str, col) -> None:
        """Selo visível quando o Claude conferiu/escolheu esta coluna."""
        if res and col and res.get(papel) == col:
            st.caption("🤖 Conferido pelo Claude")

    if _ai_v_res:
        auto_sp = _ai_pick(_ai_v_res, "telefone", df_sales_raw, auto_sp)
        auto_sn = _ai_pick(_ai_v_res, "nome", df_sales_raw, auto_sn)
        auto_vl = _ai_pick(_ai_v_res, "valor", df_sales_raw, auto_vl)
        if _ai_v_res.get("observacao"):
            st.info(f"🤖 **Vendas:** {_ai_v_res['observacao']}")
    if _ai_k_res:
        auto_kp = _ai_pick(_ai_k_res, "telefone", df_kommo_raw, auto_kp)
        auto_kt = _ai_pick(_ai_k_res, "tags", df_kommo_raw, auto_kt)
        auto_kn = _ai_pick(_ai_k_res, "nome", df_kommo_raw, auto_kn)
        if _ai_k_res.get("observacao"):
            st.info(f"🤖 **Kommo:** {_ai_k_res['observacao']}")

    def _idx(df, col):
        cols = list(df.columns)
        return cols.index(col) if col in cols else 0

    # Campos de palavra-chave com estado — permite que a IA os preencha
    for _k, _v in (("traffic_kw", "trafego"), ("traffic_exclude", ""),
                   ("disparo_kw", "disparo"), ("disparo_exclude", "")):
        if _k not in st.session_state:
            st.session_state[_k] = _v

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        sales_phone_col = st.selectbox(
            "Coluna de telefone — Vendas",
            list(df_sales_raw.columns),
            index=_idx(df_sales_raw, auto_sp),
            help="Coluna da planilha de vendas com o telefone/WhatsApp do cliente. "
                 "A comparação casa números com DDD, sem DDD, com ou sem o 9 — mas exige "
                 "que o DDD bata quando os dois lados têm, evitando juntar cidades diferentes.",
        )
        st.caption(_phone_preview(df_sales_raw, sales_phone_col, auto_sp == sales_phone_col))
        _ai_conferido(_ai_v_res, "telefone", sales_phone_col)
    with c2:
        kommo_phone_col = st.selectbox(
            "Coluna de telefone — Kommo",
            list(df_kommo_raw.columns),
            index=_idx(df_kommo_raw, auto_kp),
            help="Coluna do Kommo com o telefone do lead (geralmente 'Celular'). "
                 "Mesmo que o número esteja em outra coluna, a ferramenta testa todas automaticamente.",
        )
        st.caption(_phone_preview(df_kommo_raw, kommo_phone_col, auto_kp == kommo_phone_col))
        _ai_conferido(_ai_k_res, "telefone", kommo_phone_col)
    with c3:
        kommo_tag_col = st.selectbox(
            "Coluna de tags — Kommo",
            list(df_kommo_raw.columns),
            index=_idx(df_kommo_raw, auto_kt) if auto_kt else 0,
            help="Coluna do Kommo onde estão as etiquetas dos leads (geralmente 'Tags'). "
                 "É aqui que a ferramenta procura as palavras-chave de tráfego e de disparo.",
        )
        if auto_kt == kommo_tag_col:
            st.caption("✨ Auto-detectado")
        _ai_conferido(_ai_k_res, "tags", kommo_tag_col)
    with c4:
        traffic_keyword = st.text_input(
            "Tag de tráfego — palavra-chave",
            key="traffic_kw",
            placeholder="Cole aqui como está no Kommo...",
            help="Trecho da tag de tráfego como aparece no Kommo. Ignora maiúsculas e acentos. "
                 "Pode pôr VÁRIAS separadas por vírgula (qualquer uma serve): 'trafego, pago, ads'.",
        )
        traffic_exclude = st.text_input(
            "Excluir tags com *(opcional)*",
            placeholder="ex: organico",
            help="Se preencher, leads cuja tag contém estas palavras NÃO contam como tráfego. "
                 "Útil pra separar 'Tráfego Pago' de 'Tráfego Orgânico'. Várias separadas por vírgula.",
            key="traffic_exclude",
        )

    # ── 🤖 IA: sugestão de tags (opcional) ────────────────────────────────────
    def _tag_counts_json(df: Optional[pd.DataFrame], col: Optional[str], top: int = 60) -> dict:
        if df is None or not col or col not in df.columns:
            return {}
        vc = (df[col].fillna("").astype(str).str.split(",").explode().str.strip()
              .loc[lambda s: s != ""].value_counts().head(top))
        return {str(k): int(v) for k, v in vc.items()}

    def _cb_sugerir_tags(payload: str):
        res, err = _ai_safe(ai_sugerir_tags, payload)
        if err:
            st.session_state["_ai_tags_note"] = ("erro", err)
            return
        for campo, chave in (("trafego_incluir", "traffic_kw"),
                             ("trafego_excluir", "traffic_exclude"),
                             ("disparo_incluir", "disparo_kw"),
                             ("disparo_excluir", "disparo_exclude")):
            val = (res.get(campo) or "").strip()
            if campo.endswith("incluir") and not val:
                continue  # não apaga a palavra-chave se a IA não achou o grupo
            st.session_state[chave] = val
        st.session_state["_ai_tags_note"] = ("ok", res.get("explicacao", "Tags preenchidas."))

    _disp_tag_auto = detect_tag_col(df_kommo_disparo_raw) if df_kommo_disparo_raw is not None else None
    _tags_payload = json.dumps({
        "tags_do_kommo": _tag_counts_json(df_kommo_raw, kommo_tag_col),
        "tags_do_kommo_disparo": _tag_counts_json(df_kommo_disparo_raw, _disp_tag_auto),
    }, ensure_ascii=False)
    st.button(
        "🤖 Sugerir tags com IA",
        disabled=not _has_ai,
        on_click=_cb_sugerir_tags, args=(_tags_payload,),
        help="O Claude lê as tags dos seus leads e preenche os campos de "
             "tráfego e de disparo (incluindo exclusões) automaticamente."
             + ("" if _has_ai else " Para ativar, cole sua chave da Anthropic na barra lateral."),
    )
    _ai_tags_note = st.session_state.get("_ai_tags_note")
    if _ai_tags_note:
        (st.warning if _ai_tags_note[0] == "erro" else st.info)(f"🤖 {_ai_tags_note[1]}")

    # Colunas de nome (fallback quando telefone não bate)
    none_val = "(não usar)"
    nm1, nm2, nm3 = st.columns(3)
    with nm1:
        name_opts_s = [none_val] + list(df_sales_raw.columns)
        name_default_s = name_opts_s.index(auto_sn) if auto_sn and auto_sn in name_opts_s else 0
        sales_name_col_sel = st.selectbox(
            "Coluna de nome — Vendas *(fallback)*",
            name_opts_s, index=name_default_s, key="sales_name_col",
            help="Quando o telefone não bate, a ferramenta tenta cruzar pelo nome completo. "
                 "Selecione a coluna de nome na planilha de vendas. "
                 "Matches por nome são sinalizados com ⚠️ para revisão manual.",
        )
        sales_name_col: Optional[str] = None if sales_name_col_sel == none_val else sales_name_col_sel
        if auto_sn == sales_name_col:
            st.caption("✨ Auto-detectado")
        _ai_conferido(_ai_v_res, "nome", sales_name_col)
    with nm2:
        name_opts_k = [none_val] + list(df_kommo_raw.columns)
        name_default_k = name_opts_k.index(auto_kn) if auto_kn and auto_kn in name_opts_k else 0
        kommo_name_col_sel = st.selectbox(
            "Coluna de nome — Kommo *(fallback)*",
            name_opts_k, index=name_default_k, key="kommo_name_col",
            help="Coluna de nome dos leads no Kommo. Usada como critério de fallback.",
        )
        kommo_name_col: Optional[str] = None if kommo_name_col_sel == none_val else kommo_name_col_sel
        if auto_kn == kommo_name_col:
            st.caption("✨ Auto-detectado")
        _ai_conferido(_ai_k_res, "nome", kommo_name_col)
    with nm3:
        val_opts = [none_val] + list(df_sales_raw.columns)
        val_default = val_opts.index(auto_vl) if auto_vl and auto_vl in val_opts else 0
        sales_value_col_sel = st.selectbox(
            "Coluna de valor da venda *(opcional)*",
            val_opts, index=val_default, key="sales_value_col",
            help="Se a planilha de vendas tiver uma coluna com o valor em R$ de cada venda, selecione aqui. "
                 "A ferramenta vai calcular e exibir a receita total de tráfego e de disparo.",
        )
        sales_value_col = None if sales_value_col_sel == none_val else sales_value_col_sel
        if auto_vl == sales_value_col:
            st.caption("✨ Auto-detectado")
        _ai_conferido(_ai_v_res, "valor", sales_value_col)

    # O PROCV é telefone-first: todas as colunas de telefone dos DOIS lados são
    # varridas (cliente com 3 números casa por qualquer um; dedup por venda).
    # Nome é último recurso — e aqui o usuário pode desligá-lo de vez.
    usar_nome_fallback = st.checkbox(
        "Aceitar match por NOME quando nenhum telefone bater (entra marcado com ⚠️)",
        value=False,
        help="O cruzamento bate telefone com telefone: um cliente com 3 números cadastrados "
             "casa por qualquer um deles, em qualquer coluna — inclusive número escondido "
             "em campo de e-mail/observação. DESLIGADO por padrão para um resultado 100% "
             "por telefone. Ligando, o nome completo só casa se tiver ao menos uma palavra "
             "distintiva ('Maria José' sozinho não conta) e entra marcado com ⚠️.",
    )
    sales_name_col_match: Optional[str] = sales_name_col if usar_nome_fallback else None


    # ── Configuração de Disparo ────────────────────────────────────────────────
    st.markdown("##### 📣 Disparo")
    none_opt = "(não usar)"

    # Decide qual Kommo usar para disparo
    df_kommo_disp = df_kommo_disparo_raw if df_kommo_disparo_raw is not None else df_kommo_raw
    if df_kommo_disparo_raw is not None:
        st.caption("📣 Usando o arquivo Kommo Disparo separado para análise de disparo.")

    auto_kd = detect_date_col(df_kommo_disp)
    auto_sd = detect_date_col(df_sales_raw)
    if _ai_v_res and _ai_v_res.get("data") in df_sales_raw.columns:
        auto_sd = _ai_v_res["data"]  # IA conferiu a coluna de data da venda

    # Se há segundo Kommo, deixa o usuário escolher as colunas dele
    if df_kommo_disparo_raw is not None:
        auto_kdp = detect_phone_col(df_kommo_disparo_raw)
        auto_kdt = detect_tag_col(df_kommo_disparo_raw)
        dd1, dd2 = st.columns(2)
        with dd1:
            kommo_disp_phone_col = st.selectbox(
                "Telefone — Kommo Disparo",
                list(df_kommo_disparo_raw.columns),
                index=_idx(df_kommo_disparo_raw, auto_kdp),
                key="kommo_disp_phone",
            )
            if auto_kdp == kommo_disp_phone_col:
                st.caption("✨ Auto-detectado")
        with dd2:
            kommo_disp_tag_col = st.selectbox(
                "Coluna de tags — Kommo Disparo",
                list(df_kommo_disparo_raw.columns),
                index=_idx(df_kommo_disparo_raw, auto_kdt) if auto_kdt else 0,
                key="kommo_disp_tag",
            )
            if auto_kdt == kommo_disp_tag_col:
                st.caption("✨ Auto-detectado")
    else:
        kommo_disp_phone_col = kommo_phone_col
        kommo_disp_tag_col = kommo_tag_col

    e1, e2, e3 = st.columns(3)
    with e1:
        disparo_keyword = st.text_input(
            "Tag de disparo — palavra-chave",
            placeholder="Cole aqui como está a tag no Kommo...",
            help="Trecho da tag de disparo como aparece no Kommo. Ignora maiúsculas e acentos. "
                 "Pode pôr várias separadas por vírgula (qualquer uma serve). "
                 "Ex.: 'disparo', 'recebeu disparo', 'dia das mães'.",
            key="disparo_kw",
        )
        disparo_exclude = st.text_input(
            "Excluir tags com *(opcional)*",
            placeholder="ex: teste",
            help="Leads cuja tag contém estas palavras NÃO contam como disparo. Várias por vírgula.",
            key="disparo_exclude",
        )
    with e2:
        kommo_date_opts = [none_opt] + list(df_kommo_disp.columns)
        kommo_date_default = (kommo_date_opts.index(auto_kd)
                              if auto_kd and auto_kd in kommo_date_opts else 0)
        kommo_date_sel = st.selectbox(
            "Data do disparo no Kommo *(opcional)*",
            kommo_date_opts,
            index=kommo_date_default,
            key="kommo_date",
            help="Selecione a coluna do Kommo que contém a data em que o disparo foi feito. "
                 "Se a data estiver embutida no texto da tag (ex: 'Disparo 22/04'), "
                 "a ferramenta extrai automaticamente — deixe em '(não usar)' nesse caso. "
                 "A data do disparo é usada para garantir que só vendas APÓS o disparo sejam contadas.",
        )
        kommo_date_col = None if kommo_date_sel == none_opt else kommo_date_sel
        if auto_kd == kommo_date_col:
            st.caption("✨ Auto-detectado")
    with e3:
        sales_date_opts = [none_opt] + list(df_sales_raw.columns)
        sales_date_default = (sales_date_opts.index(auto_sd)
                              if auto_sd and auto_sd in sales_date_opts else 0)
        sales_date_sel = st.selectbox(
            "Data da venda *(opcional)*",
            sales_date_opts,
            index=sales_date_default,
            key="sales_date",
            help="Coluna da planilha de vendas com a data em que a compra foi realizada. "
                 "Usada para validar que a venda aconteceu DEPOIS do disparo (janela de 30 dias). "
                 "Se não tiver essa coluna, a ferramenta ainda encontra as vendas — "
                 "mas não consegue garantir a ordem cronológica.",
        )
        sales_date_col = None if sales_date_sel == none_opt else sales_date_sel
        if auto_sd == sales_date_col:
            st.caption("✨ Auto-detectado")
        _ai_conferido(_ai_v_res, "data", sales_date_col)

    if not kommo_date_col:
        st.caption("ℹ️ Sem coluna de data do disparo: a ferramenta tentará extrair a data do texto da tag automaticamente (ex: 'Disparo 22/04').")

    considerar_disparo = st.checkbox(
        "📣 Considerar disparo também",
        value=True,
        help="Quando marcado, analisa conversões via disparo mesmo que já existam vendas pelo tráfego.",
    )

    st.divider()

    # ── Passo 3: Processar ─────────────────────────────────────────────────────
    st.markdown('<div class="step-wrap"><div class="step-num">3</div><div class="step-text">Processar</div></div>', unsafe_allow_html=True)

    if st.button("✨  RODAR PROC AURE", use_container_width=True, type="primary"):
        progress = st.progress(0, text="Iniciando...")
        try:
            progress.progress(10, text="Cruzando vendas com Kommo...")
            ds_t, dk_t, df_result, df_full = run_procv(
                df_sales_raw, sales_phone_col,
                df_kommo_raw, kommo_phone_col, kommo_tag_col,
                traffic_keyword,
                sales_name_col=sales_name_col_match,
                kommo_name_col=kommo_name_col,
                traffic_exclude=traffic_exclude,
            )

            # Compradores únicos (deduplicado por Tel_8dig no run_procv)
            confirmed = len(df_result)

            progress.progress(60, text="Analisando disparo...")
            df_disparo_result = None
            if considerar_disparo and disparo_keyword.strip():
                df_disparo_result = run_disparo(
                    df_sales_raw, sales_phone_col, sales_date_col,
                    df_kommo_disp, kommo_disp_phone_col, kommo_disp_tag_col,
                    disparo_keyword, kommo_date_col,
                    sales_name_col=sales_name_col_match,
                    kommo_name_col=kommo_name_col,
                    disparo_exclude=disparo_exclude,
                )

            # ── Breakdown mês a mês (quando a planilha de vendas tem várias abas) ──
            # Em try próprio: se falhar/demais, o relatório principal ainda sai.
            df_breakdown = None
            if "_Planilha" in df_sales_raw.columns:
                progress.progress(70, text="Cruzando o Kommo com cada aba/mês...")
                try:
                    df_breakdown = run_breakdown_by_sheet(
                        df_sales_raw, sales_phone_col, df_kommo_raw, kommo_phone_col, kommo_tag_col,
                        traffic_keyword,
                        sales_date_col=sales_date_col,
                        disparo_keyword=disparo_keyword if (considerar_disparo and disparo_keyword.strip()) else None,
                        kommo_date_col=kommo_date_col,
                        sales_name_col=sales_name_col_match, kommo_name_col=kommo_name_col,
                        df_kommo_disp=df_kommo_disparo_raw,
                        kommo_disp_phone_col=kommo_disp_phone_col,
                        kommo_disp_tag_col=kommo_disp_tag_col,
                        traffic_exclude=traffic_exclude,
                        disparo_exclude=disparo_exclude,
                    )
                except Exception:
                    df_breakdown = None  # não bloqueia o relatório principal

            # ── Atribuição: sobreposição tráfego × disparo ────────────────────
            trafego_phones: set = set()
            if len(df_result) > 0 and "Tel_8dig" in df_result.columns:
                trafego_phones = {v for v in df_result["Tel_8dig"].dropna() if v}

            disparo_phones: set = set()
            if df_disparo_result is not None and len(df_disparo_result) > 0:
                disp_sim = df_disparo_result[df_disparo_result["Venda_Confirmada"] == "SIM"]
                if "Tel_8dig" in disp_sim.columns:
                    disparo_phones = {v for v in disp_sim["Tel_8dig"].dropna() if v}

            overlap_phones = trafego_phones & disparo_phones

            # Adiciona coluna "Origem" em cada resultado
            if len(df_result) > 0:
                df_result["Origem"] = df_result["Tel_8dig"].apply(
                    lambda t: "Tráfego + Disparo" if t in overlap_phones else "Tráfego"
                )
            if df_disparo_result is not None and len(df_disparo_result) > 0:
                df_disparo_result["Origem"] = df_disparo_result["Tel_8dig"].apply(
                    lambda t: "Tráfego + Disparo" if t in overlap_phones else "Disparo"
                )

            disp_sim_df = None
            if df_disparo_result is not None and len(df_disparo_result) > 0:
                disp_sim_df = df_disparo_result[df_disparo_result["Venda_Confirmada"] == "SIM"]

            progress.progress(80, text="Analisando duplicatas e multi-compras...")
            # Duplicatas: só em abas primárias (score ≥ 50) para evitar falsos positivos
            # de telefones que aparecem em abas de cadastro/auxiliares
            df_for_dup = df_sales_raw
            if "_Score_Aba" in df_sales_raw.columns:
                df_primary = df_sales_raw[
                    pd.to_numeric(df_sales_raw["_Score_Aba"], errors="coerce").fillna(0) >= 50
                ]
                if len(df_primary) >= 2:
                    df_for_dup = df_primary
            df_dup_analysis = analyze_duplicates(df_for_dup, sales_phone_col, sales_date_col)

            progress.progress(90, text="Gerando relatório Excel...")
            excel_bytes = build_excel(ds_t, dk_t, df_result, df_full, df_disparo_result, df_dup_analysis,
                                       sales_value_col=sales_value_col, df_breakdown=df_breakdown)
            st.session_state["excel_bytes"] = excel_bytes

            progress.progress(100, text="Concluído!")
            progress.empty()

            total_traffic = int((df_full["É_Tráfego"] == "SIM").sum())
            disp_conv  = 0
            disp_total = 0
            if df_disparo_result is not None and len(df_disparo_result) > 0:
                disp_conv  = int((df_disparo_result["Venda_Confirmada"] == "SIM").sum())
                disp_total = len(df_disparo_result)

            n_overlap    = len(overlap_phones)
            n_total_uniq = len(trafego_phones | disparo_phones)

            # Qualidade dos dados de vendas — telefone em QUALQUER coluna conta
            # (cliente com fixo vazio mas celular preenchido TEM telefone)
            n_sales_total  = len(df_sales_raw)
            _sales_tel_cols = [c for c in df_sales_raw.columns
                               if not _is_doc_col(c) and not str(c).startswith("_")
                               and _is_phone_col(df_sales_raw[c].dropna())]
            if sales_phone_col not in _sales_tel_cols:
                _sales_tel_cols.insert(0, sales_phone_col)

            def _tel_da_linha(row) -> str:
                for _c in _sales_tel_cols:
                    _v = row.get(_c)
                    if pd.notna(_v):
                        _s8 = right8(clean_phone(str(_v)))
                        if _s8:
                            return _s8
                return ""

            phones_series = df_sales_raw.apply(_tel_da_linha, axis=1)
            n_sem_tel = int((phones_series == "").sum())
            n_com_tel = n_sales_total - n_sem_tel

            # ── Resumo ─────────────────────────────────────────────────────────
            st.markdown('<div class="step-wrap"><div class="step-num">✓</div><div class="step-text">Resultados</div></div>', unsafe_allow_html=True)

            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Vendas carregadas", f"{n_sales_total:,}")
            r2.metric("Com telefone válido", f"{n_com_tel:,}",
                      delta=f"{n_sem_tel:,} sem tel." if n_sem_tel > 0 else None,
                      delta_color="off")
            r3.metric("Leads no Kommo", f"{len(dk_t):,}")
            r4.metric("Leads de tráfego no Kommo", f"{total_traffic:,}")

            if n_sem_tel > 0:
                pct = n_sem_tel / n_sales_total * 100
                st.warning(
                    f"⚠️ **{n_sem_tel} de {n_sales_total} vendas ({pct:.0f}%) não têm telefone preenchido** "
                    f"e não podem ser cruzadas com o Kommo. Os resultados abaixo refletem apenas as {n_com_tel} vendas com telefone."
                )

            st.divider()

            # ── Tráfego ────────────────────────────────────────────────────────
            st.markdown('<div class="step-wrap"><div class="step-num">📊</div><div class="step-text">Tráfego Pago</div></div>', unsafe_allow_html=True)
            if confirmed > 0:
                taxa_t = f"{confirmed/total_traffic*100:.1f}%" if total_traffic > 0 else "—"
                _por_nome = int(df_result["Criterio_Match"].astype(str)
                                .str.contains("Nome", na=False).sum())
                _por_tel = confirmed - _por_nome
                _detalhe = (f" · **{_por_tel} por telefone**"
                            + (f" + **{_por_nome} por nome ⚠️**" if _por_nome else ""))
                st.success(f"**{confirmed} vendas** encontradas de tráfego pago — {taxa_t} "
                           f"de conversão sobre {total_traffic:,} leads{_detalhe}.")
                st.caption("Veja a lista completa na aba **✅ Vendas — Tráfego** do Excel.")
                with st.expander(f"Prévia — {confirmed} vendas de tráfego"):
                    st.dataframe(df_result, use_container_width=True, height=260)
            else:
                # ── Quando 0 tráfego: ajuda o usuário a entender o que existe ──
                all_phone_matches = df_full[df_full["Venda_Confirmada"] == "SIM"]
                n_any_match = len(all_phone_matches)

                if n_any_match > 0:
                    st.warning(
                        f"Nenhuma venda encontrada com a tag **\"{traffic_keyword}\"**. "
                        f"Porém, **{n_any_match} leads do Kommo têm telefone que bate com suas vendas** — "
                        f"pode ser que a tag de tráfego use um nome diferente."
                    )
                    # Top tags no Kommo para o usuário identificar a tag certa
                    top_tags = (
                        df_kommo_raw[kommo_tag_col]
                        .fillna("").astype(str)
                        .str.split(",")
                        .explode()
                        .str.strip()
                        .loc[lambda s: s != ""]
                        .value_counts()
                        .head(12)
                    )
                    if len(top_tags) > 0:
                        with st.expander("💡 Tags mais usadas no Kommo — qual identifica o tráfego pago?"):
                            st.caption("Cole uma dessas tags no campo 'Tag de tráfego' acima e rode novamente:")
                            for tag, cnt in top_tags.items():
                                st.markdown(f"• **{tag}** — {cnt:,} leads")

                    with st.expander(f"📋 {n_any_match} leads com telefone encontrado nas vendas (sem filtro de tag)"):
                        st.caption(
                            "Esses leads do Kommo têm telefone que bate com uma venda. "
                            "Veja a coluna 'Tag_Kommo' para identificar quais são do tráfego pago."
                        )
                        st.dataframe(all_phone_matches, use_container_width=True, height=300)
                else:
                    st.info(
                        f"Nenhuma venda encontrada com a tag **\"{traffic_keyword}\"** e "
                        f"nenhum telefone do Kommo bateu com as vendas. "
                        f"Verifique se a coluna de telefone está correta nos dois arquivos."
                    )

            # ── Disparo ────────────────────────────────────────────────────────
            if considerar_disparo and disparo_keyword.strip():
                st.divider()
                st.markdown('<div class="step-wrap"><div class="step-num">📣</div><div class="step-text">Disparo (WhatsApp)</div></div>', unsafe_allow_html=True)

                if df_disparo_result is None or disp_total == 0:
                    # Sugere tags alternativas
                    top_tags_d = (
                        df_kommo_disp[kommo_disp_tag_col]
                        .fillna("").astype(str)
                        .str.split(",")
                        .explode()
                        .str.strip()
                        .loc[lambda s: s != ""]
                        .value_counts()
                        .head(12)
                    )
                    st.warning(f"Nenhum lead com a tag **\"{disparo_keyword}\"** encontrado no Kommo.")
                    if len(top_tags_d) > 0:
                        with st.expander("💡 Tags disponíveis no Kommo"):
                            st.caption("Cole uma dessas no campo 'Tag de disparo' acima:")
                            for tag, cnt in top_tags_d.items():
                                st.markdown(f"• **{tag}** — {cnt:,} leads")
                elif disp_conv > 0:
                    taxa_d = f"{disp_conv/disp_total*100:.1f}%"
                    st.success(f"**{disp_conv} vendas** encontradas de disparo — {taxa_d} de conversão sobre {disp_total:,} leads disparados.")
                    if not (kommo_date_col or sales_date_col):
                        st.caption("Sem datas configuradas — todos os matches de telefone foram incluídos. Filtre por Data_Venda no Excel.")
                    st.caption("Veja a lista completa na aba **📣 Vendas — Disparo** do Excel.")
                    with st.expander(f"Prévia — {disp_conv} vendas de disparo"):
                        st.dataframe(
                            df_disparo_result[df_disparo_result["Venda_Confirmada"] == "SIM"],
                            use_container_width=True, height=260,
                        )
                else:
                    st.warning(
                        f"**{disp_total:,} leads de disparo** encontrados, mas nenhuma venda confirmada. "
                        f"Possíveis causas: as vendas ocorreram antes do disparo, fora da janela de 30 dias, "
                        f"ou os telefones não bateram."
                    )

            # ── Sobreposição ────────────────────────────────────────────────────
            if n_overlap > 0:
                st.divider()
                st.info(
                    f"🔀 **{n_overlap} comprador(es)** aparecem em tráfego **e** disparo. "
                    f"Total único de compradores: **{n_total_uniq}**. "
                    f"Detalhes na coluna **'Origem da Venda'** do Excel."
                )

            # ── 🤖 Contexto do resultado (alimenta a análise E o chat) ─────────
            _crit_traf = (df_result["Criterio_Match"].value_counts().to_dict()
                          if len(df_result) > 0 else {})
            _datas_venda = None
            if sales_date_col:
                _dts = [d for d in df_sales_raw[sales_date_col].head(2000).apply(parse_date)
                        if _is_real_datetime(d)]
                if _dts:
                    _datas_venda = {"de": min(_dts).strftime("%d/%m/%Y"),
                                    "ate": max(_dts).strftime("%d/%m/%Y")}
            _datas_disp: dict = {}
            if considerar_disparo and disparo_keyword.strip():
                _mask_d = df_kommo_disp[kommo_disp_tag_col].fillna("").astype(str).apply(
                    lambda t: _tag_matches(t, disparo_keyword, disparo_exclude))
                for _t in df_kommo_disp.loc[_mask_d, kommo_disp_tag_col].astype(str).head(800):
                    for _seg in _t.split(","):
                        if _tag_matches(_seg, disparo_keyword, disparo_exclude):
                            _dd = parse_date(_seg)
                            if _is_real_datetime(_dd):
                                _k = _dd.strftime("%d/%m/%Y")
                                _datas_disp[_k] = _datas_disp.get(_k, 0) + 1
            _resumo_dict = {
                "vendas": {"total": int(n_sales_total), "com_telefone": int(n_com_tel),
                           "sem_telefone": int(n_sem_tel), "periodo": _datas_venda},
                "kommo": {"leads": int(len(dk_t)),
                          "top_tags": _tag_counts_json(df_kommo_raw, kommo_tag_col, 12)},
                "trafego": {"palavra_chave": traffic_keyword, "excluir": traffic_exclude,
                            "leads_com_tag": int(total_traffic),
                            "conversoes": int(confirmed),
                            "criterios_de_match": {k: int(v) for k, v in _crit_traf.items()},
                            "bloqueados_por_ddd_divergente": int(df_full.attrs.get("n_ddd_blocked", 0))},
                "disparo": {"palavra_chave": disparo_keyword,
                            "analisado": bool(considerar_disparo and disparo_keyword.strip()),
                            "leads_de_disparo": int(disp_total),
                            "conversoes_na_janela_30d": int(disp_conv),
                            "datas_dos_disparos_extraidas_das_tags": _datas_disp},
                "sobreposicao_trafego_e_disparo": int(n_overlap),
            }
            _payload_resumo = json.dumps(_resumo_dict, ensure_ascii=False, default=str)

            def _tbl_chat(df: Optional[pd.DataFrame], n: int = 150) -> list:
                """Tabela compacta das conversões para o chat 🤖."""
                if df is None or len(df) == 0:
                    return []
                cols = [c for c in ("Origem", "Tag_Kommo", "Nome_Kommo", "Tel_8dig",
                                    "Criterio_Match", "Data_Venda", "Dias_Após_Disparo")
                        if c in df.columns]
                for _c in (sales_name_col, sales_value_col, sales_date_col):
                    if _c and f"[Venda] {_c}" in df.columns:
                        cols.append(f"[Venda] {_c}")
                return df[cols].head(n).astype(str).to_dict("records")

            # Colunas de telefone das vendas: preenchimento e validade de cada uma
            _tel_cols_stats = {}
            for _c in df_sales_raw.columns:
                if str(_c).startswith("_"):
                    continue
                _cd = df_sales_raw[_c].dropna()
                if len(_cd) > 0 and _is_phone_col(_cd):
                    _tel_cols_stats[str(_c)] = {
                        "preenchidos": int(len(_cd)),
                        "telefones_validos": int(_cd.apply(
                            lambda v: phone_key(v)[1] != "").sum()),
                    }

            _nao_conv = df_full[(df_full["É_Tráfego"] == "SIM")
                                & (df_full["Venda_Confirmada"] == "NÃO")]
            _nao_conv_amostra = [
                {"nome": str(_r.get("Nome_Kommo", "")),
                 "telefone": _mask_pii(str(_r.get("Telefone_Kommo", ""))),
                 "tags": str(_r.get("Tag_Kommo", ""))[:60]}
                for _, _r in _nao_conv.head(20).iterrows()
            ]
            _sem_tel_amostra = []
            if sales_name_col:
                _sem_tel_amostra = (df_sales_raw.loc[phones_series == "", sales_name_col]
                                    .dropna().astype(str).head(10).tolist())

            st.session_state["_ai_ctx"] = json.dumps({
                "resumo": _resumo_dict,
                "tratamento_aplicado_nas_planilhas": {
                    "vendas": list(df_sales_raw.attrs.get("treatment_notes", [])),
                    "kommo": list(df_kommo_raw.attrs.get("treatment_notes", [])),
                },
                "colunas": {
                    "todas_da_planilha_de_vendas": [str(c) for c in df_sales_raw.columns][:80],
                    "colunas_de_telefone_das_vendas": _tel_cols_stats,
                    "mapeamento_usado": {
                        "telefone_vendas": sales_phone_col, "nome_vendas": sales_name_col,
                        "valor_venda": sales_value_col, "data_venda": sales_date_col,
                        "telefone_kommo": kommo_phone_col, "tags_kommo": kommo_tag_col,
                        "nome_kommo": kommo_name_col, "data_disparo_kommo": kommo_date_col,
                    },
                    "observacoes_da_ia_sobre_as_planilhas": {
                        "vendas": (_ai_v_res or {}).get("observacao"),
                        "kommo": (_ai_k_res or {}).get("observacao"),
                    },
                },
                "como_a_ferramenta_cruza": (
                    "Varre TODAS as colunas com cara de telefone dos dois lados (não só a "
                    "principal) e RESGATA números escondidos em outras colunas (ex.: telefone "
                    "no campo de e-mail com prefixo 'p:+55'), nos dois lados; normaliza "
                    "com/sem DDD, com/sem o 9, DDI 55, pontos/traços; o match exige DDD igual "
                    "quando os dois lados têm DDD; fallback por nome é OPCIONAL (desligado "
                    "por padrão), exige 2+ palavras com ao menos uma distintiva ('Maria José' "
                    "sozinho não casa) e entra marcado com ⚠️; deduplica por comprador "
                    "(telefone) e por venda; disparo: vale a venda de 0 a 30 dias após a "
                    "data do disparo (extraída da tag ou da coluna)."),
                "conversoes_de_trafego": _tbl_chat(df_result),
                "conversoes_de_disparo": _tbl_chat(disp_sim_df),
                "resultado_por_lista_ou_mes": (df_breakdown.astype(str).to_dict("records")
                                              if (df_breakdown is not None and len(df_breakdown) > 0) else []),
                "leads_de_trafego_que_NAO_converteram_amostra": _nao_conv_amostra,
                "vendas_sem_telefone_amostra": _sem_tel_amostra,
            }, ensure_ascii=False, default=str)
            # Tabelas REAIS ficam na sessão — o chat consulta linha a linha via tool
            st.session_state["_ai_dfs"] = {
                "vendas": df_sales_raw,
                "kommo": df_kommo_raw,
                "conversoes_trafego": df_result,
                "todos_os_leads_cruzados": df_full,
                "disparo": df_disparo_result,
            }
            st.session_state.pop("_ai_chat", None)  # novo processamento → conversa nova

            # ── 🤖 Análise do Claude ────────────────────────────────────────────
            if not _has_ai:
                st.divider()
                st.caption("🤖 Ative a **Inteligência Claude** na barra lateral para receber aqui "
                           "a análise executiva, a revisão dos matches por nome e o chat sobre o resultado.")
            if _has_ai and ai_auto:
                st.divider()
                st.markdown('<div class="step-wrap"><div class="step-num">🤖</div><div class="step-text">Análise do Claude</div></div>', unsafe_allow_html=True)

                # 1) Revisão dos matches ⚠️ por nome (tráfego + disparo, até 30)
                _pares = []
                if sales_name_col:
                    _fontes = [("tráfego", df_result)]
                    if disp_sim_df is not None:
                        _fontes.append(("disparo", disp_sim_df))
                    for _origem, _df_f in _fontes:
                        if _df_f is None or len(_df_f) == 0:
                            continue
                        for _, _r in _df_f.iterrows():
                            if "Nome" not in str(_r.get("Criterio_Match", "")):
                                continue
                            _pares.append({
                                "indice": len(_pares),
                                "nome_no_kommo": str(_r.get("Nome_Kommo", "")),
                                "nome_na_venda": str(_r.get(f"[Venda] {sales_name_col}", "")),
                                "origem": _origem,
                            })
                            if len(_pares) >= 30:
                                break
                        if len(_pares) >= 30:
                            break
                if _pares:
                    with st.spinner(f"🤖 Revisando {len(_pares)} matches por nome..."):
                        _rev, _rev_err = _ai_safe(
                            ai_revisar_matches,
                            json.dumps({"pares": _pares}, ensure_ascii=False))
                    if _rev_err:
                        st.warning(f"🤖 Revisão de matches indisponível: {_rev_err}")
                    elif _rev:
                        _emoji = {"mesma_pessoa": "✅", "incerto": "🤔",
                                  "pessoas_diferentes": "❌"}
                        _vlist = _rev.get("veredictos", [])
                        _n_susp = sum(1 for v in _vlist
                                      if v.get("veredito") != "mesma_pessoa")
                        _titulo = f"🤖 Revisão dos {len(_pares)} matches por nome (⚠️)"
                        if _n_susp:
                            _titulo += f" — {_n_susp} merecem conferência"
                        with st.expander(_titulo, expanded=_n_susp > 0):
                            for v in _vlist:
                                try:
                                    p = _pares[int(v.get("indice", -1))]
                                except (IndexError, ValueError, TypeError):
                                    continue
                                st.markdown(
                                    f"{_emoji.get(v.get('veredito'), '❓')} "
                                    f"**{p['nome_no_kommo']}** × **{p['nome_na_venda']}** "
                                    f"({p['origem']}) — {v.get('motivo', '')}"
                                )
                            st.caption("Veredicto do Claude sobre os matches sem telefone batendo. "
                                       "Confirme os ❌ e 🤔 antes de fechar o relatório.")

                # 2) Resumo executivo dos números (payload montado acima)
                with st.spinner("🤖 Gerando análise executiva..."):
                    _txt, _txt_err = _ai_safe(ai_explicar_resultados, _payload_resumo)
                if _txt_err:
                    st.caption(f"🤖 Análise executiva indisponível: {_txt_err}")
                elif _txt:
                    st.markdown(_txt)

            # ── Mês a mês (quando a planilha de vendas tem várias abas) ──────────
            if df_breakdown is not None and len(df_breakdown) > 0:
                st.divider()
                st.markdown('<div class="step-wrap"><div class="step-num">📅</div><div class="step-text">Por Lista / Mês (cada arquivo ou aba)</div></div>', unsafe_allow_html=True)
                st.caption(
                    "Cada aba da planilha de vendas foi cruzada com o Kommo **separadamente**. "
                    "Um comprador recorrente conta em **cada mês** que comprou — diferente do total "
                    "acima, que conta compradores únicos no período."
                )
                st.dataframe(df_breakdown, use_container_width=True, hide_index=True)
                _tot_t = int(df_breakdown["Vendas de Tráfego"].sum())
                _msg = f"**{_tot_t}** vendas de tráfego somando os meses"
                if "Vendas de Disparo" in df_breakdown.columns:
                    _msg += f" · **{int(df_breakdown['Vendas de Disparo'].sum())}** de disparo"
                st.success(_msg + ". Veja a aba **📅 Mês a Mês** do Excel.")

            # ── Qualidade dos dados de vendas ───────────────────────────────────
            st.divider()
            st.markdown('<div class="step-wrap"><div class="step-num">🔍</div><div class="step-text">Qualidade dos Dados</div></div>', unsafe_allow_html=True)

            sit      = df_dup_analysis["Situacao_Venda"]
            n_dup    = int(sit.str.startswith("DUPLICATA").sum())
            n_multi  = int(sit.str.startswith("Multi").sum())

            qd1, qd2, qd3, qd4 = st.columns(4)
            qd1.metric("Total de vendas", f"{n_sales_total:,}")
            qd2.metric("Com telefone", f"{n_com_tel:,}")
            qd3.metric("Sem telefone", f"{n_sem_tel:,}", delta_color="off")
            qd4.metric("Duplicatas suspeitas", f"{n_dup:,}")

            if n_dup > 0:
                st.error(f"⚠️ **{n_dup} registros duplicados** — mesmo número, data e conteúdo idêntico. Veja aba **🔍 Duplicatas** no Excel.")
            if n_multi > 0:
                st.caption(f"🔁 {n_multi} multi-compras identificadas (mesmo cliente, datas diferentes) — normal para clientes recorrentes. Ver aba 🔍 Duplicatas no Excel.")
            if n_dup == 0 and n_multi == 0:
                st.success("Nenhuma duplicata detectada nos dados de vendas.")

            # ── O que o cruzamento tratou (telefones soltos + falsos matches evitados) ──
            n_stray = df_full.attrs.get("n_stray_sales", 0)
            n_blocked = df_full.attrs.get("n_ddd_blocked", 0)
            if n_stray or n_blocked:
                msgs = []
                if n_stray:
                    msgs.append(
                        f"📌 **{n_stray} telefone(s) fora da coluna de telefone** foram "
                        f"localizados e cruzados mesmo assim (marcados com ⚠️ no Excel)."
                    )
                if n_blocked:
                    msgs.append(
                        f"🛡️ **{n_blocked} falso(s) match(es) evitado(s)** — números com o mesmo "
                        f"final mas DDD de cidade diferente **não** foram contados como a mesma pessoa."
                    )
                st.info("🧠 **O que o cruzamento tratou:**\n\n" + "\n\n".join(msgs))

            st.divider()

        except Exception as e:
            progress.empty()
            st.error(
                "⚠️ Não consegui concluir a análise. As causas mais comuns são: "
                "coluna de **telefone** ou de **tags** apontada errada, ou uma planilha "
                "num formato inesperado. Confira as colunas selecionadas acima e tente de novo."
            )
            st.caption(f"Detalhe técnico: {type(e).__name__}: {e}")
            with st.expander("🔧 Detalhes técnicos (para suporte)"):
                st.code(traceback.format_exc())

    # ── Download — fora do bloco de processamento para persistir entre reruns ──
    if "excel_bytes" in st.session_state:
        st.divider()
        st.download_button(
            label="📥  Baixar Excel — Resultado Completo",
            data=st.session_state["excel_bytes"],
            file_name="proc_aure_resultado.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption(
            "💡 Para usar no Google Sheets: faça upload do .xlsx no Google Drive → "
            "clique com botão direito → Abrir com → Planilhas Google."
        )

else:
    st.info("⬆️ Carregue as duas planilhas acima para continuar.")

# ── 💬 Chat com o Claude sobre o resultado ─────────────────────────────────────
# Fora do bloco de processamento: o contexto fica na sessão e a conversa
# sobrevive aos reruns (cada mensagem do chat re-executa o script).
if st.session_state.get("_ai_ctx"):
    st.divider()
    st.markdown('<div class="step-wrap"><div class="step-num">💬</div><div class="step-text">Converse com o Claude sobre este resultado</div></div>', unsafe_allow_html=True)
    try:
        _ctx_resumo = json.loads(st.session_state["_ai_ctx"]).get("resumo", {})
        st.caption(
            f"Contexto carregado: **{_ctx_resumo.get('vendas', {}).get('total', '?')} vendas** · "
            f"**{_ctx_resumo.get('trafego', {}).get('conversoes', '?')} conversões de tráfego** · "
            f"**{_ctx_resumo.get('disparo', {}).get('conversoes_na_janela_30d', '?')} de disparo**. "
            "O Claude supervisiona ESTE processamento com acesso às tabelas reais — "
            "ele consulta as planilhas linha a linha para responder. Pergunte, questione, corrija."
        )
    except Exception:
        pass
    if not _has_ai:
        st.caption("🤖 Cole sua chave da Anthropic na barra lateral para conversar sobre o resultado.")
    else:
        _hist = st.session_state.setdefault("_ai_chat", [])
        _licoes = [ln.strip() for m in _hist if m["role"] == "assistant"
                   for ln in m["content"].split("\n") if ln.strip().startswith("📌")]
        if _licoes:
            with st.expander(f"📌 {len(_licoes)} lição(ões) que o Claude anotou nesta conversa"):
                for ln in _licoes:
                    st.markdown(f"- {ln.removeprefix('📌').strip()}")
        for _m in _hist:
            with st.chat_message(_m["role"], avatar=("🧑‍💼" if _m["role"] == "user" else "🤖")):
                st.markdown(_m["content"])
        _pergunta = st.chat_input("Ex.: por que a venda da Maria não apareceu no tráfego?")
        if _pergunta:
            _hist.append({"role": "user", "content": _pergunta})
            with st.chat_message("user", avatar="🧑‍💼"):
                st.markdown(_pergunta)
            with st.chat_message("assistant", avatar="🤖"):
                with st.spinner("🤖 Analisando o resultado..."):
                    _resp_chat, _err_chat = _ai_chat_reply(st.session_state["_ai_ctx"], _hist)
                if _err_chat:
                    st.warning(f"🤖 {_err_chat}")
                    _hist.pop()  # devolve a pergunta — um erro não trava a conversa
                else:
                    st.markdown(_resp_chat)
                    _hist.append({"role": "assistant", "content": _resp_chat})

# ── Rodapé ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<div class="footer">Desenvolvido por <b>João</b> &nbsp;·&nbsp; Proc Aure &nbsp;·&nbsp; Aure Digital</div>',
    unsafe_allow_html=True,
)
